"""
Extrai os indicadores das 7 IEES paranaenses + 33 IES nacionais de comparação (total: 40 IES).

Uso:
    python pipeline/assemble_final.py

Saída:
    - stdout: JSON completo (indicadores + fontes)
    - data/seti_precomputed.json: JSON para o dashboard (atualizado automaticamente)
    - stderr: tabela de resumo das 7 IES PR
"""
import json
import sys
import datetime
import unicodedata
import openpyxl
from pathlib import Path

# Garante UTF-8 no stdout (Windows usa cp1252 por padrão)
sys.stdout.reconfigure(encoding="utf-8")

DATA_DIR = Path(__file__).parent.parent / "data"

# IES paranaenses — têm dados do Relatório 8050, Suplementação, Clusterização, CBO2/RAIS
IEES_PR = ["UEL", "UEM", "UEPG", "UNIOESTE", "UNICENTRO", "UENP", "UNESPAR"]

# IES nacionais de comparação — têm dados de Cursos, IES, CAPES (bases Brasil)
IEES_BR = [
    # 15 originais
    "USP", "UNESP", "UNICAMP", "UERJ", "UDESC", "UERGS",
    "UECE", "UNEB", "UESB", "UEG", "UEMA", "UEPB", "UEPA", "UEA", "UERN",
    # 17 novas + URCA (substitui UNIVESP, que é 100% EaD e incomparável)
    "UESC", "UNCISAL", "UVA", "UNIMONTES", "UPE", "UEFS", "UNEMAT",
    "UESPI", "UNITINS", "UENF", "UEMS", "UEMG", "UERR", "UNEAL",
    "UEAP", "UEMASUL", "UnDF", "URCA",
]

IEES = IEES_PR + IEES_BR

CO_IES_MAP = {
    # Paraná
    9:     "UEL",
    57:    "UEM",
    730:   "UEPG",
    609:   "UNIOESTE",
    1126:  "UNICENTRO",
    15015: "UENP",
    18492: "UNESPAR",
    # Nacionais — 15 originais (CO confirmados via Base IES - Brasil.xlsx)
    55:    "USP",      # CO=55 confirmado
    56:    "UNESP",    # CO=56 confirmado (era 55 incorretamente)
    54:    "UNICAMP",
    547:   "UERJ",
    43:    "UDESC",
    3336:  "UERGS",
    29:    "UECE",
    40:    "UNEB",
    688:   "UESB",
    47:    "UEG",
    568:   "UEMA",
    550:   "UEPB",
    38:    "UEPA",
    3172:  "UEA",
    71:    "UERN",
    # Novas 17 IES estaduais
    24:    "UESC",
    32:    "UNCISAL",
    95:    "UVA",
    367:   "UNIMONTES",
    409:   "UPE",
    666:   "UEFS",
    719:   "UNEMAT",
    756:   "UESPI",
    829:   "UNITINS",
    1027:  "UENF",
    1028:  "UEMS",
    1036:  "UEMG",
    5077:  "UERR",
    5242:  "UNEAL",
    5701:  "UEAP",
    23410: "UEMASUL",
    27103: "UnDF",
    746:   "URCA",
}

INDICATORS = [
    "students", "entrants", "graduates", "courses", "vacancies",
    "occupancy", "dropout", "completion", "doctors",
    "cnpq", "capes", "pg", "pgTop",
    # Seção 5 — Base_Cursos (CAPES): breakdown por grau, área e município
    "pgMestrado", "pgMestradoProf", "pgDoutorado",
    "pgPorGrandeArea", "pgMunicipiosDistintos",
    # Seção 5b — Base_Discentes (CAPES)
    "discMestrado", "discDoutorado",
    "tituladosMestrado", "tituladosDoutorado",
    "pctExcelencia",
    # Seção 5b — Base_Docentes (CAPES)
    "docPermanentes", "docColaboradores", "docVisitantes",
    "razaoDocenteDiscente",
    "budget", "execution", "liquidation", "personnel", "supplementation",
    "employment", "salary", "insertionRatePR",
    "facultyOcc", "cres", "tide",
    "docVagasTotais", "docVagasDisp", "docVagasOcupadas", "docTaxaUtil",
    "docVagasCond", "docPctCond", "docTideAtrib", "docTidePartic",
    "docTidePctNaoAtrib", "docChMedia", "docCresAut", "docCresUtil",
    "docCresSaldo", "docCresOciosidade", "docCresPartic",
    "egressosMunicipios",
    "seloNotaFinal",
    # Seção 12d — V9 (Estratificação_IES_Estaduais_BR.xlsx): filtro de
    # UNIVERSO de IES por Grande Área CINE predominante. Não é variável de
    # groupBy (V1-V8) nem entra em clusters_raw.
    "areaCineGrande", "areaCinePct", "areaCineHerfindahl",
]


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_pct(v, already_pct=False):
    if v is None:
        return None
    try:
        f = float(v)
        if already_pct:
            return round(f, 2)
        return round(f * 100, 2) if f <= 1.0 else round(f, 2)
    except Exception:
        return None


def safe_float(v, d=2):
    if v is None:
        return None
    try:
        return round(float(v), d)
    except Exception:
        return None


def safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _blank():
    return {k: None for k in INDICATORS}


# ── inicialização ─────────────────────────────────────────────────────────────

results = {iees.lower(): _blank() for iees in IEES}
sources = {iees.lower(): {} for iees in IEES}

_pipeline_start = datetime.datetime.now()
_pipeline_alerts: list[str] = []


# ── 1. INEP — Proporção de docentes com doutorado ────────────────────────────
# Fonte: Base IES - Brasil.xlsx / Base_ IES_BRASIL
# Coluna: "Proporção de docentes com doutorado" × 100
# Ano: mais recente por IES (NU_ANO_CENSO)

wb = openpyxl.load_workbook(DATA_DIR / "Base IES - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["Base_ IES_BRASIL"]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
col_idx = {h: i for i, h in enumerate(headers) if h is not None}
co_col = col_idx.get("CO_IES")
yr_col = col_idx.get("NU_ANO_CENSO")

ies_raw = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    co = row[co_col]
    try:
        co = int(co)
    except Exception:
        continue
    if co not in CO_IES_MAP:
        continue
    y = row[yr_col]
    try:
        y = int(y)
    except Exception:
        y = 0
    iees = CO_IES_MAP[co]
    if iees not in ies_raw or y > ies_raw[iees]["_year"]:
        ies_raw[iees] = {"_year": y}
        for k, i in col_idx.items():
            ies_raw[iees][k] = row[i]
wb.close()

for iees in IEES:
    key = iees.lower()
    if iees not in ies_raw:
        continue
    for cn, v in ies_raw[iees].items():
        if cn and "doutorado" in str(cn).lower() and "propor" in str(cn).lower() and v is not None:
            results[key]["doctors"] = round(float(v) * 100, 2)
            sources[key]["doctors"] = (
                f"Base IES - Brasil.xlsx / Base_ IES_BRASIL"
                f" / '{cn}' × 100 / ano={ies_raw[iees]['_year']}"
            )
            break


# ── 2. INEP — Matrículas, ingressos, concluintes, cursos, vagas ──────────────
# Fonte: Base Cursos - Brasil.xlsx / _IES PÚBLICAS ESTADUAIS_CURSOS
# Colunas: QT_MAT, QT_ING, QT_CONC, QT_VG_TOTAL, QT_CURSO, QT_SIT_DESVINCULADO
# Transformação: soma por IES no ano mais recente;
#   occupancy = QT_ING / QT_VG_TOTAL × 100 (ingressantes / vagas ofertadas)
#   dropout   = QT_SIT_DESVINCULADO / QT_MAT × 100
#   completion = QT_CONC / QT_MAT × 100

wb = openpyxl.load_workbook(DATA_DIR / "Base Cursos - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["_IES PÚBLICAS ESTADUAIS_CURSOS "]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
col_idx = {h: i for i, h in enumerate(headers) if h is not None}

yr_col  = col_idx.get("NU_ANO_CENSO")
co_col  = col_idx.get("CO_IES")
qt_mat  = col_idx.get("QT_MAT")
qt_ing  = col_idx.get("QT_ING")
qt_conc = col_idx.get("QT_CONC")
qt_vg   = col_idx.get("QT_VG_TOTAL")
qt_curs = col_idx.get("QT_CURSO")
qt_desv = col_idx.get("QT_SIT_DESVINCULADO")

# Round 4 — Tipo de Vaga / Turno de Vaga (filtro global, seletor de métrica):
# colunas confirmadas na Etapa 1 do diagnóstico (todas somam para QT_VG_TOTAL/
# QT_ING, ver diagnóstico). QT_VG_PROC_SELETIVO / QT_ING_PROC_SELETIVO existem
# no cabeçalho mas são 0 em 100% das linhas nesta base (2020-2024) — mantidos
# mesmo assim para não esconder o campo caso uma base futura venha preenchida.
# Não existe QT_ING_EAD (ingressantes por turno EaD) — occupancy de Turno=EaD
# fica sempre null (contagem bruta), decisão confirmada com a Luíza.
qt_vg_diurno    = col_idx.get("QT_VG_TOTAL_DIURNO")
qt_vg_noturno   = col_idx.get("QT_VG_TOTAL_NOTURNO")
qt_vg_ead       = col_idx.get("QT_VG_TOTAL_EAD")
qt_vg_nova      = col_idx.get("QT_VG_NOVA")
qt_vg_procsel   = col_idx.get("QT_VG_PROC_SELETIVO")
qt_vg_remanesc  = col_idx.get("QT_VG_REMANESC")
qt_vg_progesp   = col_idx.get("QT_VG_PROG_ESPECIAL")
qt_ing_diurno   = col_idx.get("QT_ING_DIURNO")
qt_ing_noturno  = col_idx.get("QT_ING_NOTURNO")
qt_ing_vgnova   = col_idx.get("QT_ING_VG_NOVA")
qt_ing_procsel  = col_idx.get("QT_ING_PROC_SELETIVO")
qt_ing_remanesc = col_idx.get("QT_ING_VG_REMANESC")
qt_ing_progesp  = col_idx.get("QT_ING_VG_PROG_ESPECIAL")

cursos_data = {}
# Round 3a: cursos_data_by_year retém TODAS as linhas de cada (IES, ano), em
# paralelo a cursos_data (que só guarda o ano mais recente por IES, como
# antes) — mesma passada única sobre a planilha, sem reabrir o arquivo.
# Alimenta cursos_detalhado_by_year (Seção 2b) para permitir Tipo de
# Curso/Modalidade/Grande Área combinados com o filtro de Ano (Round 3b).
cursos_data_by_year = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    co = row[co_col]
    try:
        co = int(co)
    except Exception:
        continue
    if co not in CO_IES_MAP:
        continue
    y = row[yr_col]
    try:
        y = int(y)
    except Exception:
        y = 0
    iees = CO_IES_MAP[co]
    if iees not in cursos_data:
        cursos_data[iees] = {"_year": 0, "_rows": []}
    if y >= cursos_data[iees]["_year"]:
        if y > cursos_data[iees]["_year"]:
            cursos_data[iees] = {"_year": y, "_rows": []}
        cursos_data[iees]["_rows"].append(row)
    cursos_data_by_year.setdefault(iees, {}).setdefault(y, []).append(row)
wb.close()

for iees in IEES:
    key = iees.lower()
    if iees not in cursos_data:
        continue
    rows = cursos_data[iees]["_rows"]
    year = cursos_data[iees]["_year"]

    students  = sum(safe_int(r[qt_mat])  or 0 for r in rows)
    entrants  = sum(safe_int(r[qt_ing])  or 0 for r in rows)
    graduates = sum(safe_int(r[qt_conc]) or 0 for r in rows)
    vacancies = sum(safe_int(r[qt_vg])   or 0 for r in rows)
    courses   = sum(safe_int(r[qt_curs]) or 0 for r in rows)
    desvinc   = sum(safe_int(r[qt_desv]) or 0 for r in rows)

    occupancy  = round(entrants  / vacancies * 100, 2) if vacancies > 0 else None
    completion = round(graduates / students  * 100, 2) if students  > 0 else None
    dropout    = round(desvinc   / students  * 100, 2) if students  > 0 else None

    results[key].update(
        students=students, entrants=entrants, graduates=graduates,
        vacancies=vacancies, courses=courses,
        occupancy=occupancy, completion=completion, dropout=dropout,
    )
    src = f"Base Cursos - Brasil.xlsx / _IES PÚBLICAS ESTADUAIS_CURSOS / ano={year} / soma por CO_IES"
    sources[key].update(
        students=src + " / QT_MAT",
        entrants=src + " / QT_ING",
        graduates=src + " / QT_CONC",
        vacancies=src + " / QT_VG_TOTAL",
        courses=src + " / QT_CURSO",
        occupancy=src + " / QT_ING ÷ QT_VG_TOTAL × 100",
        dropout=src + " / QT_SIT_DESVINCULADO ÷ QT_MAT × 100",
        completion=src + " / QT_CONC ÷ QT_MAT × 100",
    )


# ── 2b. Cursos — desagregação por Grande Área CINE, Grau Acadêmico e Modalidade
# Fonte: mesmas linhas já coletadas em cursos_data (Seção 2) — não reabre o arquivo.
# Chave de agregação: (NO_CINE_AREA_GERAL, TP_GRAU_ACADEMICO, TP_MODALIDADE_ENSINO)
#
# ATENÇÃO: os rótulos de TP_GRAU_ACADEMICO e TP_MODALIDADE_ENSINO seguem o padrão
# do dicionário do Censo da Educação Superior/INEP, mas NÃO foram confirmados
# contra o dicionário oficial anexo ao Censo neste projeto — [CONFIRMAR NO CÓDIGO]
# antes de expor ao usuário final. Valor 0 e None em TP_GRAU_ACADEMICO ocorrem em
# cursos interdisciplinares (ex.: "Abi - Ciências Biológicas") e ficam como
# "Não classificado" em vez de assumir um rótulo.

_GRAU_LABELS = {
    1: "Bacharelado",
    2: "Licenciatura",
    3: "Tecnólogo",     # confirmado pela Luíza — rótulo correto
    0: "Não classificado",
    None: "Não classificado",
}
_MODALIDADE_LABELS = {
    1: "Presencial",  # [CONFIRMAR NO CÓDIGO — dicionário INEP]
    2: "EaD",         # [CONFIRMAR NO CÓDIGO — dicionário INEP]
}

cine_col  = col_idx.get("NO_CINE_AREA_GERAL")
grau_col  = col_idx.get("TP_GRAU_ACADEMICO")
modal_col = col_idx.get("TP_MODALIDADE_ENSINO")

# Round 3a: lógica de agrupamento extraída para função — usada tanto por
# cursos_detalhado (ano mais recente, comportamento inalterado) quanto por
# cursos_detalhado_by_year (todos os anos, novo) para garantir fórmulas
# idênticas entre os dois, sem duplicar a lógica.
def _agrupar_cursos_por_grupo(_rows_d):
    _grupos = {}
    for _r in _rows_d:
        _cine = _r[cine_col] if cine_col is not None else None
        try:
            _grau_int = int(_r[grau_col]) if grau_col is not None and _r[grau_col] is not None else None
        except Exception:
            _grau_int = None
        try:
            _modal_int = int(_r[modal_col]) if modal_col is not None and _r[modal_col] is not None else None
        except Exception:
            _modal_int = None
        _chave = (
            _cine or "Não informado",
            _GRAU_LABELS.get(_grau_int, "Não classificado"),
            _MODALIDADE_LABELS.get(_modal_int, "Não informado"),
        )
        if _chave not in _grupos:
            _grupos[_chave] = {
                "mat": 0, "ing": 0, "conc": 0, "vgTot": 0, "cursos": 0, "desvinc": 0,
                # Round 4 — Tipo de Vaga / Turno (contagens brutas por grupo;
                # o front-end soma essas contagens entre grupos já restringidos
                # por Tipo de Curso/Modalidade/Grande Área e SÓ ENTÃO calcula a
                # razão — mesmo padrão de "dropoutCount" acima, já que % não é
                # somável entre grupos).
                "vgDiurno": 0, "vgNoturno": 0, "vgEad": 0,
                "vgNova": 0, "vgProcSeletivo": 0, "vgRemanesc": 0, "vgProgEspecial": 0,
                "ingDiurno": 0, "ingNoturno": 0,
                "ingVgNova": 0, "ingProcSeletivo": 0, "ingVgRemanesc": 0, "ingVgProgEspecial": 0,
            }
        _g = _grupos[_chave]
        _g["mat"]     += safe_int(_r[qt_mat])  or 0
        _g["ing"]     += safe_int(_r[qt_ing])  or 0
        _g["conc"]    += safe_int(_r[qt_conc]) or 0
        _g["vgTot"]   += safe_int(_r[qt_vg])   or 0
        _g["cursos"]  += safe_int(_r[qt_curs]) or 0
        _g["desvinc"] += safe_int(_r[qt_desv]) or 0
        _g["vgDiurno"]         += safe_int(_r[qt_vg_diurno])   or 0
        _g["vgNoturno"]        += safe_int(_r[qt_vg_noturno])  or 0
        _g["vgEad"]            += safe_int(_r[qt_vg_ead])      or 0
        _g["vgNova"]           += safe_int(_r[qt_vg_nova])     or 0
        _g["vgProcSeletivo"]   += safe_int(_r[qt_vg_procsel])  or 0
        _g["vgRemanesc"]       += safe_int(_r[qt_vg_remanesc]) or 0
        _g["vgProgEspecial"]   += safe_int(_r[qt_vg_progesp])  or 0
        _g["ingDiurno"]        += safe_int(_r[qt_ing_diurno])   or 0
        _g["ingNoturno"]       += safe_int(_r[qt_ing_noturno])  or 0
        _g["ingVgNova"]        += safe_int(_r[qt_ing_vgnova])   or 0
        _g["ingProcSeletivo"]  += safe_int(_r[qt_ing_procsel])  or 0
        _g["ingVgRemanesc"]    += safe_int(_r[qt_ing_remanesc]) or 0
        _g["ingVgProgEspecial"] += safe_int(_r[qt_ing_progesp]) or 0

    _lista = []
    for (_cine, _grau_label, _modal_label), _g in _grupos.items():
        _occ  = round(_g["ing"]  / _g["vgTot"] * 100, 2) if _g["vgTot"] > 0 else None
        _comp = round(_g["conc"] / _g["mat"]   * 100, 2) if _g["mat"]   > 0 else None
        _drop = round(_g["desvinc"] / _g["mat"] * 100, 2) if _g["mat"]  > 0 else None
        _lista.append({
            "cineArea": _cine,
            "grauAcademico": _grau_label,
            "modalidade": _modal_label,
            "students": _g["mat"],
            "entrants": _g["ing"],
            "graduates": _g["conc"],
            "vacancies": _g["vgTot"],
            "courses": _g["cursos"],
            "occupancy": _occ,
            "completion": _comp,
            "dropout": _drop,
            # Contagem bruta de QT_SIT_DESVINCULADO do grupo — exposta para que o
            # frontend possa recalcular "dropout" ao somar grupos (ex.: filtro de
            # Tipo de Curso, Round 2b), já que "dropout" acima é uma razão e não
            # é somável entre grupos. Mesma fórmula: dropoutCount / students * 100.
            "dropoutCount": _g["desvinc"],
            # Round 4 — Tipo de Vaga / Turno: contagens brutas (vagas e
            # ingressantes correspondentes) por grupo. Ver comentário do
            # dict de acumulação acima sobre por que não expomos % pronto aqui.
            "vgDiurno": _g["vgDiurno"], "vgNoturno": _g["vgNoturno"], "vgEad": _g["vgEad"],
            "vgNova": _g["vgNova"], "vgProcSeletivo": _g["vgProcSeletivo"],
            "vgRemanesc": _g["vgRemanesc"], "vgProgEspecial": _g["vgProgEspecial"],
            "ingDiurno": _g["ingDiurno"], "ingNoturno": _g["ingNoturno"],
            "ingVgNova": _g["ingVgNova"], "ingProcSeletivo": _g["ingProcSeletivo"],
            "ingVgRemanesc": _g["ingVgRemanesc"], "ingVgProgEspecial": _g["ingVgProgEspecial"],
        })
    return _lista


cursos_detalhado = {}
for _iees_d, _info_d in cursos_data.items():
    cursos_detalhado[_iees_d] = _agrupar_cursos_por_grupo(_info_d["_rows"])

# Round 3a: cursos_detalhado_by_year — {sigla: {str(ano): [grupos]}}, mesmo
# padrão de d8050_by_year (Seção 9). Pré-requisito para o front-end (Round
# 3b) permitir Tipo de Curso/Modalidade/Grande Área combinados com o filtro
# de Ano. cursos_detalhado (acima) continua sendo calculado e exportado sem
# nenhuma alteração — os dois convivem em paralelo nesta rodada.
cursos_detalhado_by_year = {}
for _iees_d, _por_ano in cursos_data_by_year.items():
    cursos_detalhado_by_year[_iees_d] = {
        str(_ano): _agrupar_cursos_por_grupo(_linhas)
        for _ano, _linhas in _por_ano.items()
    }


# ── 2c. Benchmark por Grande Área CINE — referência PR vs Brasil ─────────────
# Fonte: mesma cursos_detalhado computada acima (Seção 2b) — não reabre a base.
# Recorte "pr": as 7 IES-PR (IEES_PR). Recorte "br": as 40 IES (IEES = IEES_PR +
# IEES_BR) — mesma definição de escopo Brasil usada no restante do painel
# (ex. scopeUniverse() no frontend, que também é IES-PR + IES-BR, não IES-BR
# isolada).
#
# DECISÕES NÃO VALIDADAS COM JÉSSICA/ANDERSON (pendência de validação
# metodológica — ver changelog):
#   - Ponderação: cada indicador pondera pelo seu próprio denominador de taxa —
#     dropout pondera por "students" (mat), occupancyTipo pondera por
#     "vacancies" (vgTot). Não há campo único "vagas ou matrículas" comum aos
#     dois; a escolha acima segue a semântica de cada fórmula, não foi pedida
#     explicitamente.
#   - "occupancyTipo" reaproveita o campo "occupancy" já existente em
#     cursos_detalhado, que é ingressantes ÷ vagas (QT_ING/QT_VG_TOTAL) — igual
#     ao já usado pelos getters ind26/ind67 em painel.js. A fórmula oficial do
#     IND-67 no catálogo é matrículas ÷ vagas (QT_MAT/QT_VG_TOTAL); não existe
#     em cursos_detalhado nenhum campo com essa fórmula alternativa, e não foi
#     criado um novo campo para isso nesta rodada.
#   - NOTA 2026-07-13: tentativa de correção para matrículas÷vagas revertida
#     — resultado produziu valores >100% (estoque plurianual ÷ fluxo anual
#     de vagas). Fórmula correta pendente de validação com SETI. Ver
#     README.md para detalhes.
#   - Áreas "Não informado" (CINE ausente/nulo) entram como mais uma área se
#     aparecerem nos dados, igual ao comportamento já existente em
#     cineAreaOptions() no frontend (painel-aba3-acesso.js) — não filtradas.

_POLARIDADE_CINE = {
    "dropout":       "menor",  # menor valor vence (taxa de desvinculação)
    "occupancyTipo": "maior",  # maior valor vence (taxa de ocupação)
}
_CAMPO_CINE = {
    "dropout":       "dropout",
    "occupancyTipo": "occupancy",
}
_PESO_CINE = {
    "dropout":       "students",
    "occupancyTipo": "vacancies",
}


# Round 3a: fonte migrada de cursos_detalhado (flat, ano mais recente) para
# cursos_detalhado_by_year, selecionando explicitamente o ano mais recente
# por IES — mesmo critério que cursos_data já usava (max(y) por IES).
# Resultado idêntico ao anterior (confirmado: cursos_detalhado[iees] ==
# cursos_detalhado_by_year[iees][str(ano_mais_recente)], mesmas linhas de
# origem); o benchmark CINE continua fora do escopo de "filtrar por ano"
# (fotografia fixa), só a fonte de leitura mudou para não depender mais do
# campo antigo, que o Round 3b pode vir a aposentar no front-end.
def _grupos_ano_mais_recente_cine(iees):
    anos = cursos_detalhado_by_year.get(iees)
    if not anos:
        return []
    ano_recente = max(anos.keys(), key=lambda a: int(a))
    return anos[ano_recente]


def _media_ponderada_cine(lista_iees, area, metrica):
    campo = _CAMPO_CINE[metrica]
    peso_campo = _PESO_CINE[metrica]
    soma_valor_peso = 0.0
    soma_peso = 0.0
    for _iees_b in lista_iees:
        for _g in _grupos_ano_mais_recente_cine(_iees_b):
            if _g["cineArea"] != area:
                continue
            _v = _g[campo]
            _p = _g[peso_campo]
            if _v is None or not _p:
                continue
            soma_valor_peso += _v * _p
            soma_peso += _p
    if soma_peso <= 0:
        return None
    return round(soma_valor_peso / soma_peso, 2)


def calcular_benchmark_cine():
    areas = sorted({
        _g["cineArea"]
        for _iees_b in cursos_detalhado_by_year
        for _g in _grupos_ano_mais_recente_cine(_iees_b)
    })
    resultado = {}
    for _area in areas:
        _entry = {}
        for _metrica, _pol in _POLARIDADE_CINE.items():
            _v_pr = _media_ponderada_cine(IEES_PR, _area, _metrica)
            _v_br = _media_ponderada_cine(IEES, _area, _metrica)
            if _v_pr is None and _v_br is None:
                _ref, _origem = None, None
            elif _v_pr is None:
                _ref, _origem = _v_br, "BR"
            elif _v_br is None:
                _ref, _origem = _v_pr, "PR"
            elif _pol == "menor":
                _ref, _origem = (_v_pr, "PR") if _v_pr <= _v_br else (_v_br, "BR")
            else:
                _ref, _origem = (_v_pr, "PR") if _v_pr >= _v_br else (_v_br, "BR")
            _entry[_metrica] = {"pr": _v_pr, "br": _v_br, "referencia": _ref, "origem": _origem}
        resultado[_area] = _entry
    return resultado


benchmark_cine = calcular_benchmark_cine()


# ── 3. Docentes — facultyOcc, cres, tide ─────────────────────────────────────
# Fonte: Base Docentes - Paraná.xlsx / Base_Docentes_PR
# Colunas (0-based):
#   [0]  ANO        [2]  IEES
#   [20] Taxa de ocupação do quadro docente         → facultyOcc
#   [25] Participação do TIDE no quadro disponível  → tide  (novo)
#   [30] Taxa de utilização da CRES                 → cres
# Ano: mais recente por IES; dentro do ano, último registro do arquivo

wb = openpyxl.load_workbook(DATA_DIR / "Base Docentes - Paraná.xlsx", read_only=True, data_only=True)
ws = wb["Base_Docentes_PR"]
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

doc_latest = {}  # iees → {"_year": int, "_row": tuple}
for row in ws.iter_rows(min_row=2, values_only=True):
    iees = row[2] if len(row) > 2 else None
    if iees not in IEES_PR:
        continue
    try:
        y = int(row[0])
    except (TypeError, ValueError):
        continue
    if iees not in doc_latest or y >= doc_latest[iees]["_year"]:
        doc_latest[iees] = {"_year": y, "_row": row}
wb.close()

for iees in IEES_PR:
    key = iees.lower()
    if iees not in doc_latest:
        continue
    row = doc_latest[iees]["_row"]
    y   = doc_latest[iees]["_year"]
    total_codes = row[17] if len(row) > 17 else None
    vagas_disp  = row[18] if len(row) > 18 else None
    vagas_ocup  = row[19] if len(row) > 19 else None
    occ         = row[20] if len(row) > 20 else None
    taxa_util   = row[21] if len(row) > 21 else None
    vagas_cond  = row[22] if len(row) > 22 else None
    pct_cond    = row[23] if len(row) > 23 else None  # Excel column X / IND-49
    tide_atrib  = row[24] if len(row) > 24 else None
    tide        = row[25] if len(row) > 25 else None
    tide_nao    = row[26] if len(row) > 26 else None
    ch_media    = row[27] if len(row) > 27 else None
    cres_aut    = row[28] if len(row) > 28 else None
    cres_util   = row[29] if len(row) > 29 else None
    cres        = row[30] if len(row) > 30 else None
    cres_saldo  = row[31] if len(row) > 31 else None
    cres_ocios  = row[32] if len(row) > 32 else None
    cres_partic = row[33] if len(row) > 33 else None

    # col[20-26,30,32,33] armazenados como decimais (0.0–1.0+) na base.
    # safe_pct trata valores > 1 como "já são %", o que quebra casos como
    # UNIOESTE (taxa CRES e utilização acima de 100%). A correção é forçar ×100
    # para os campos que sabemos serem decimais, usando safe_float diretamente.
    def _dec_to_pct(v):
        try:
            return round(float(v) * 100, 2) if v is not None else None
        except Exception:
            return None

    results[key]["docVagasTotais"] = safe_int(total_codes)
    results[key]["docVagasDisp"] = safe_int(vagas_disp)
    results[key]["docVagasOcupadas"] = safe_int(vagas_ocup)
    results[key]["facultyOcc"] = _dec_to_pct(occ)
    results[key]["docTaxaUtil"] = _dec_to_pct(taxa_util)
    results[key]["docVagasCond"] = safe_int(vagas_cond)
    results[key]["docPctCond"] = _dec_to_pct(pct_cond)
    results[key]["docTideAtrib"] = safe_int(tide_atrib)
    results[key]["tide"] = _dec_to_pct(tide)
    results[key]["docTidePartic"] = _dec_to_pct(tide)
    results[key]["docTidePctNaoAtrib"] = _dec_to_pct(tide_nao)
    results[key]["docChMedia"] = safe_float(ch_media, 2)
    results[key]["docCresAut"] = safe_int(cres_aut)
    results[key]["docCresUtil"] = safe_int(cres_util)
    results[key]["cres"] = _dec_to_pct(cres)
    results[key]["docCresSaldo"] = safe_int(cres_saldo)
    results[key]["docCresOciosidade"] = _dec_to_pct(cres_ocios)
    results[key]["docCresPartic"] = _dec_to_pct(cres_partic)
    src = f"Base Docentes - Paraná.xlsx / Base_Docentes_PR / ano={y}"
    sources[key]["facultyOcc"] = src + " / Taxa de ocupação do quadro docente (col 20)"
    sources[key]["cres"]       = src + " / Taxa de utilização da CRES (col 30)"
    sources[key]["tide"]       = src + " / Participação do TIDE no quadro disponível (col 25)"
    sources[key]["docVagasTotais"] = src + " / Total de códigos de vagas docentes (col 17)"
    sources[key]["docVagasDisp"] = src + " / Vagas docentes disponíveis para ocupação (col 18)"
    sources[key]["docVagasOcupadas"] = src + " / Vagas docentes efetivas ocupadas (col 19)"
    sources[key]["docTaxaUtil"] = src + " / Taxa de utilização das vagas docentes disponíveis (col 21)"
    sources[key]["docVagasCond"] = src + " / Vagas docentes condicionadas à autorização governamental (col 22)"
    sources[key]["docPctCond"] = src + " / Percentual de vagas condicionadas à autorização governamental (col 23 / Excel X)"
    sources[key]["docTideAtrib"] = src + " / Quantidade de TIDE atribuído ao corpo docente (col 24)"
    sources[key]["docTidePartic"] = src + " / Participação do TIDE no quadro disponível (col 25)"
    sources[key]["docTidePctNaoAtrib"] = src + " / Percentual de TIDE não atribuído (col 26)"
    sources[key]["docChMedia"] = src + " / Carga horária média de docentes efetivos (col 27)"
    sources[key]["docCresAut"] = src + " / Carga horária CRES autorizada (col 28)"
    sources[key]["docCresUtil"] = src + " / Carga horária CRES utilizada (col 29)"
    sources[key]["docCresSaldo"] = src + " / Saldo de carga horária CRES não utilizada (col 31)"
    sources[key]["docCresOciosidade"] = src + " / Taxa de ociosidade da CRES (col 32)"
    sources[key]["docCresPartic"] = src + " / Participação da CRES no esforço docente total (col 33)"


# ── 4. CNPq — captação de recursos para pesquisa ─────────────────────────────
# Fonte: Base CNPq - Brasil.xlsx / Base_CNPq_BR
# Colunas: "01_Instituição", "Ano", "Captação de recursos para pesquisa"
# Transformação: soma por IES no ano mais recente (R$ milhões)
# Match por nome da instituição (busca substring).
# A base CNPq não usa acentos — normalizamos antes do match para evitar falsos
# negativos. Lambdas usam strings sem acento. "ESTAD" captura tanto "ESTADO"
# (ex: Univ. do Estado de X) quanto "ESTADUAL" (ex: Univ. Estadual de X).

def _cnpq_norm(s: str) -> str:
    return unicodedata.normalize("NFKD", s.upper()).encode("ascii", "ignore").decode("ascii")

wb = openpyxl.load_workbook(DATA_DIR / "Base CNPq - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["Base_CNPq_BR"]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
col_idx  = {h: i for i, h in enumerate(headers) if h is not None}
inst_col = col_idx.get("01_Instituição")
yr_col   = col_idx.get("Ano")
val_col  = col_idx.get("Captação de recursos para pesquisa")

CNPQ_MATCH = {
    # ── IES-PR ──────────────────────────────────────────────────────────────
    "UEL":      lambda s: "LONDRINA" in s or "UEL" in s,
    "UEM":      lambda s: "MARINGA" in s,
    "UEPG":     lambda s: "PONTA GROSSA" in s,
    "UNIOESTE": lambda s: "OESTE DO PARANA" in s or "UNIOESTE" in s,
    "UNICENTRO":lambda s: "ESTADUAL DO CENTRO-OESTE" in s,
    "UENP":     lambda s: "NORTE DO PARANA" in s,
    "UNESPAR":  lambda s: "ESTADUAL DO PARANA" in s and "NORTE" not in s and "OESTE" not in s,
    # ── IES-BR — 15 originais ────────────────────────────────────────────────
    # ⚠ Validar valores no stderr após rodar — ordem de grandeza esperada:
    #   USP ~R$100-200M; UNICAMP/UNESP ~R$50-100M; demais ~R$2-30M
    "USP":      lambda s: ("UNIVERSIDADE DE SAO PAULO" in s or " USP" in s) and "ESTADUAL" not in s,
    "UNESP":    lambda s: "PAULISTA" in s and "ESTADUAL" in s,
    "UNICAMP":  lambda s: "CAMPINAS" in s and "ESTADUAL" in s,
    "UERJ":     lambda s: "RIO DE JANEIRO" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UDESC":    lambda s: "SANTA CATARINA" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UERGS":    lambda s: "RIO GRANDE DO SUL" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UECE":     lambda s: "CEARA" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UNEB":     lambda s: "BAHIA" in s and "ESTAD" in s and "FEDERAL" not in s and "SUDOESTE" not in s and "FEIRA" not in s,
    "UESB":     lambda s: "SUDOESTE" in s and "BAHIA" in s,
    "UEG":      lambda s: "GOIAS" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UEMA":     lambda s: "MARANHAO" in s and "ESTAD" in s and "FEDERAL" not in s and "TOCANT" not in s,
    "UEPB":     lambda s: "PARAIBA" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UEPA":     lambda s: "PARA" in s and "ESTAD" in s and "FEDERAL" not in s and "OESTE" not in s and "MARANHAO" not in s,
    "UEA":      lambda s: "AMAZONAS" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UERN":     lambda s: "RIO GRANDE DO NORTE" in s and "ESTAD" in s and "FEDERAL" not in s,
    # ── IES-BR — 17 novas ───────────────────────────────────────────────────
    "UESC":     lambda s: "SANTA CRUZ" in s and "ESTAD" in s,
    "UNCISAL":  lambda s: "CIENCIAS DA SAUDE" in s and "ALAGOAS" in s,
    "UVA":      lambda s: "VALE DO ACARAU" in s or ("UVA" in s and "CEARA" in s),
    "UNIMONTES":lambda s: "MONTES CLAROS" in s,
    "UPE":      lambda s: "PERNAMBUCO" in s and "UNIVERSIDADE DE" in s and "FEDERAL" not in s,
    "UEFS":     lambda s: "FEIRA DE SANTANA" in s,
    "UNEMAT":   lambda s: "MATO GROSSO" in s and "ESTAD" in s and "SUL" not in s and "FEDERAL" not in s,
    "UESPI":    lambda s: "PIAUI" in s and "ESTAD" in s,
    "UNITINS":  lambda s: "TOCANTINS" in s and "ESTADUAL" in s,
    "UENF":     lambda s: "NORTE FLUMINENSE" in s,
    "UEMS":     lambda s: "MATO GROSSO DO SUL" in s and "ESTAD" in s and "FEDERAL" not in s,
    "UEMG":     lambda s: "MINAS GERAIS" in s and "ESTAD" in s and "FEDERAL" not in s and "MONTES" not in s,
    "UERR":     lambda s: "RORAIMA" in s and "ESTAD" in s,
    "UNEAL":    lambda s: "ALAGOAS" in s and "ESTADUAL" in s and "SAUDE" not in s,
    "UEAP":     lambda s: "AMAPA" in s and "ESTAD" in s,
    "UEMASUL":  lambda s: "TOCANTINA" in s or "UEMASUL" in s,
    "UnDF":     lambda s: "DISTRITO FEDERAL" in s and "UNIVERSIDADE" in s and "INDUSTRIA" not in s and "SENAI" not in s and "SAUDE" not in s and "SERVICO" not in s,
    "URCA":     lambda s: "CARIRI" in s and ("REGIONAL" in s or "URCA" in s),
}

cnpq_data = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    inst = row[inst_col] if inst_col is not None else None
    y    = row[yr_col]   if yr_col   is not None else None
    val  = row[val_col]  if val_col  is not None else None
    if not inst or not isinstance(inst, str):
        continue
    inst_up = _cnpq_norm(inst)
    for iees in CNPQ_MATCH:
        if CNPQ_MATCH[iees](inst_up):
            try:
                y_int = int(y)
            except Exception:
                y_int = 0
            try:
                v = float(val) if val else 0
            except Exception:
                v = 0
            cnpq_data[(iees, y_int)] = cnpq_data.get((iees, y_int), 0) + v
            break
wb.close()

for iees in IEES:
    key = iees.lower()
    years = [y for (i, y) in cnpq_data if i == iees]
    if years:
        max_y = max(years)
        total = cnpq_data.get((iees, max_y), 0)
        results[key]["cnpq"] = round(total / 1e6, 3)
        sources[key]["cnpq"] = (
            f"Base CNPq - Brasil.xlsx / Base_CNPq_BR"
            f" / Captação de recursos para pesquisa / ano={max_y} (R$ milhões)"
        )


# ── 5. CAPES — programas de pós-graduação ────────────────────────────────────
# Fonte: Base CAPES- Pós-Graduação - Brasil.xlsx / Base_Cursos
# Colunas: CD_CONCEITO_CURSO, NM_PROGRAMA_IES, CO_IES, AN_BASE,
#          "Conceito médio dos programas de pós-graduação"
# Transformação: pg = programas distintos; pgTop = programas com conceito ≥ 5;
#   capes = conceito médio (coluna pré-calculada se disponível, senão média dos conceitos)

wb = openpyxl.load_workbook(DATA_DIR / "Base CAPES- Pós-Graduação - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["Base_Cursos"]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
col_idx    = {h: i for i, h in enumerate(headers) if h is not None}
yr_col     = col_idx.get("AN_BASE")
co_col     = col_idx.get("CO_IES")
sg_col     = col_idx.get("SG_ENTIDADE_ENSINO")
prog_col   = col_idx.get("NM_PROGRAMA_IES")
conc_col   = col_idx.get("CD_CONCEITO_CURSO")
media_col  = col_idx.get("Conceito médio dos programas de pós-graduação")
grau_col   = col_idx.get("NM_GRAU_CURSO")           # em Base_Cursos o campo chama NM_GRAU_CURSO
area_col   = col_idx.get("NM_GRANDE_AREA_CONHECIMENTO")
muni_col   = col_idx.get("NM_MUNICIPIO_PROGRAMA_IES")

capes_raw = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    co = row[co_col] if co_col is not None else None
    try:
        co_int = int(co)
    except Exception:
        co_int = 0
    iees = CO_IES_MAP.get(co_int)
    if iees is None:
        sg = row[sg_col] if sg_col is not None else None
        if sg:
            sg_up = str(sg).upper()
            iees = next((i for i in IEES if i in sg_up), None)
    if iees not in IEES:
        continue

    y = row[yr_col]
    try:
        y_int = int(y)
    except Exception:
        y_int = 0
    prog  = row[prog_col]  if prog_col  is not None else None
    c     = row[conc_col]  if conc_col  is not None else None
    media = row[media_col] if media_col is not None else None
    grau  = row[grau_col]  if grau_col  is not None else None
    area  = row[area_col]  if area_col  is not None else None
    muni  = row[muni_col]  if muni_col  is not None else None
    try:
        c_int = int(c) if c is not None else 0
    except Exception:
        c_int = 0

    if iees not in capes_raw:
        capes_raw[iees] = {}
    if y_int not in capes_raw[iees]:
        capes_raw[iees][y_int] = {"progs": {}, "media": None, "graus": {}, "areas": {}, "munis": set()}
        # graus: {NM_PROGRAMA_IES: set(NM_GRAU_CURSO)} — set porque programas combinados
        # têm uma linha por grau (ex.: MESTRADO e DOUTORADO em linhas separadas)

    if media is not None and capes_raw[iees][y_int]["media"] is None:
        try:
            mv = float(media)
            if mv > 0:
                capes_raw[iees][y_int]["media"] = mv
        except Exception:
            pass

    if prog:
        cur = capes_raw[iees][y_int]["progs"].get(prog, 0)
        if c_int > cur:
            capes_raw[iees][y_int]["progs"][prog] = c_int
        if grau is not None:
            g_str = str(grau).upper().strip()
            if prog not in capes_raw[iees][y_int]["graus"]:
                capes_raw[iees][y_int]["graus"][prog] = set()
            capes_raw[iees][y_int]["graus"][prog].add(g_str)
        if area is not None:
            capes_raw[iees][y_int]["areas"][prog] = str(area).strip()
    if muni is not None:
        capes_raw[iees][y_int]["munis"].add(str(muni).strip())
wb.close()

for iees in IEES:
    key = iees.lower()
    if iees not in capes_raw:
        continue
    years = sorted(capes_raw[iees].keys(), reverse=True)
    if not years:
        continue
    max_y = years[0]
    data  = capes_raw[iees][max_y]
    progs = data["progs"]
    media = data["media"]
    all_c = [v for v in progs.values() if v > 0]

    pg_total = len(progs)
    pg_top   = sum(1 for c in all_c if c >= 5)
    capes_mean = safe_float(media) if media is not None else (
        round(sum(all_c) / len(all_c), 2) if all_c else None
    )
    results[key].update(pg=pg_total, pgTop=pg_top, capes=capes_mean)
    src = f"Base CAPES- Pós-Graduação - Brasil.xlsx / Base_Cursos / AN_BASE={max_y}"
    sources[key].update(
        pg=src + " / NM_PROGRAMA_IES distintos",
        pgTop=src + " / CD_CONCEITO_CURSO ≥ 5",
        capes=src + " / Conceito médio dos programas",
    )

    # pgMestrado / pgMestradoProf / pgDoutorado / pgPorGrandeArea / pgMunicipiosDistintos
    graus = data.get("graus", {})
    areas = data.get("areas", {})
    munis = data.get("munis", set())
    pg_mestrado      = 0
    pg_mestrado_prof = 0
    pg_doutorado     = 0
    pg_por_area: dict = {}
    for pn, g_set in graus.items():
        # g_set é set(NM_GRAU_CURSO); programas combinados têm múltiplos valores
        # e são contados em cada categoria correspondente (conforme especificação)
        if "MESTRADO" in g_set:
            pg_mestrado += 1
        if "MESTRADO PROFISSIONAL" in g_set:
            pg_mestrado_prof += 1
        if "DOUTORADO" in g_set or "DOUTORADO PROFISSIONAL" in g_set:
            pg_doutorado += 1
    for pn, av in areas.items():
        pg_por_area[av] = pg_por_area.get(av, 0) + 1
    pg_munis = len(munis)
    results[key].update(
        pgMestrado=pg_mestrado,
        pgMestradoProf=pg_mestrado_prof,
        pgDoutorado=pg_doutorado,
        pgPorGrandeArea=pg_por_area,
        pgMunicipiosDistintos=pg_munis,
    )
    sources[key].update(
        pgMestrado=src + " / NM_PROGRAMA_IES distintos com NM_GRAU_CURSO=MESTRADO (programas combinados contados em ambas as categorias)",
        pgMestradoProf=src + " / NM_PROGRAMA_IES distintos com NM_GRAU_CURSO=MESTRADO PROFISSIONAL",
        pgDoutorado=src + " / NM_PROGRAMA_IES distintos com NM_GRAU_CURSO=DOUTORADO ou DOUTORADO PROFISSIONAL (programas combinados contados em ambas as categorias)",
        pgPorGrandeArea=src + " / NM_PROGRAMA_IES distintos por NM_GRANDE_AREA_CONHECIMENTO",
        pgMunicipiosDistintos=src + " / NM_MUNICIPIO_PROGRAMA_IES distintos",
    )


# ── 5b. CAPES — Discentes e Docentes ─────────────────────────────────────────
# Fonte: Base CAPES- Pós-Graduação - Brasil.xlsx / Base_Discentes + Base_Docentes
# Ano:   mais recente disponível por IES (AN_BASE), mesmo critério da Seção 5.
# IMPORTANTE: CD_CONCEITO_PROGRAMA (conceito do programa) ≠ CD_CONCEITO_CURSO.
#   pctExcelencia usa CD_CONCEITO_PROGRAMA >= 6 — NÃO é equivalente a pgTop.
# Docentes contados via set(ID_PESSOA) por (CO_IES, AN_BASE, DS_CATEGORIA_DOCENTE).

wb5b = openpyxl.load_workbook(
    DATA_DIR / "Base CAPES- Pós-Graduação - Brasil.xlsx",
    read_only=True, data_only=True,
)

# ── 5b.1 Base_Discentes ──────────────────────────────────────────────────────
ws_disc = wb5b["Base_Discentes"]
h_disc  = list(next(ws_disc.iter_rows(min_row=1, max_row=1, values_only=True)))
ci_disc = {h: i for i, h in enumerate(h_disc) if h is not None}
d_yr    = ci_disc.get("AN_BASE")
d_co    = ci_disc.get("CO_IES")
d_sg    = ci_disc.get("SG_ENTIDADE_ENSINO")
d_pid   = ci_disc.get("ID_PESSOA")
d_grau  = ci_disc.get("DS_GRAU_ACADEMICO_DISCENTE")
d_sit   = ci_disc.get("NM_SITUACAO_DISCENTE")
d_prog  = ci_disc.get("NM_PROGRAMA_IES")
d_conc  = ci_disc.get("CD_CONCEITO_PROGRAMA")

# disc_raw[iees][ano] = {
#   "mat_m": set(ID_PESSOA) — matriculados mestrado
#   "mat_d": set(ID_PESSOA) — matriculados doutorado
#   "tit_m": set(ID_PESSOA) — titulados mestrado
#   "tit_d": set(ID_PESSOA) — titulados doutorado
#   "progs_conc": {NM_PROGRAMA_IES: max(CD_CONCEITO_PROGRAMA)} → pctExcelencia
# }
disc_raw = {}

for row in ws_disc.iter_rows(min_row=2, values_only=True):
    co = row[d_co] if d_co is not None else None
    try:
        co_int = int(co)
    except Exception:
        co_int = 0
    iees = CO_IES_MAP.get(co_int)
    if iees is None:
        sg = row[d_sg] if d_sg is not None else None
        if sg:
            iees = next((i for i in IEES if i in str(sg).upper()), None)
    if iees not in IEES:
        continue

    y = row[d_yr] if d_yr is not None else None
    try:
        y_int = int(y)
    except Exception:
        y_int = 0

    pid  = row[d_pid]  if d_pid  is not None else None
    grau = row[d_grau] if d_grau is not None else None
    sit  = row[d_sit]  if d_sit  is not None else None
    prog = row[d_prog] if d_prog is not None else None
    conc = row[d_conc] if d_conc is not None else None

    grau_up = str(grau).upper().strip() if grau else ""
    sit_up  = str(sit).upper().strip()  if sit  else ""

    if iees not in disc_raw:
        disc_raw[iees] = {}
    if y_int not in disc_raw[iees]:
        disc_raw[iees][y_int] = {
            "mat_m": set(), "mat_d": set(),
            "tit_m": set(), "tit_d": set(),
            "progs_conc": {},
        }
    bucket = disc_raw[iees][y_int]

    if pid is not None:
        if sit_up == "MATRICULADO":
            if grau_up == "MESTRADO":
                bucket["mat_m"].add(pid)
            elif grau_up == "DOUTORADO":
                bucket["mat_d"].add(pid)
        elif sit_up == "TITULADO":
            if grau_up == "MESTRADO":
                bucket["tit_m"].add(pid)
            elif grau_up == "DOUTORADO":
                bucket["tit_d"].add(pid)

    if prog is not None and conc is not None:
        try:
            conc_int = int(conc)
        except Exception:
            conc_int = 0
        cur_conc = bucket["progs_conc"].get(prog, 0)
        if conc_int > cur_conc:
            bucket["progs_conc"][prog] = conc_int

# ── 5b.2 Base_Docentes ──────────────────────────────────────────────────────
ws_doc  = wb5b["Base_Docentes"]
h_doc   = list(next(ws_doc.iter_rows(min_row=1, max_row=1, values_only=True)))
ci_doc  = {h: i for i, h in enumerate(h_doc) if h is not None}
dc_yr   = ci_doc.get("AN_BASE")
dc_co   = ci_doc.get("CO_IES")
dc_sg   = ci_doc.get("SG_ENTIDADE_ENSINO")
dc_pid  = ci_doc.get("ID_PESSOA")
dc_cat  = ci_doc.get("DS_CATEGORIA_DOCENTE")

# doc_raw[iees][ano] = {
#   "permanente":  set(ID_PESSOA)
#   "colaborador": set(ID_PESSOA)
#   "visitante":   set(ID_PESSOA)
# }
doc_raw = {}

for row in ws_doc.iter_rows(min_row=2, values_only=True):
    co = row[dc_co] if dc_co is not None else None
    try:
        co_int = int(co)
    except Exception:
        co_int = 0
    iees = CO_IES_MAP.get(co_int)
    if iees is None:
        sg = row[dc_sg] if dc_sg is not None else None
        if sg:
            iees = next((i for i in IEES if i in str(sg).upper()), None)
    if iees not in IEES:
        continue

    y = row[dc_yr] if dc_yr is not None else None
    try:
        y_int = int(y)
    except Exception:
        y_int = 0

    pid    = row[dc_pid] if dc_pid is not None else None
    cat    = row[dc_cat] if dc_cat is not None else None
    cat_up = str(cat).upper().strip() if cat else ""

    if iees not in doc_raw:
        doc_raw[iees] = {}
    if y_int not in doc_raw[iees]:
        doc_raw[iees][y_int] = {"permanente": set(), "colaborador": set(), "visitante": set()}
    if pid is not None:
        bk = doc_raw[iees][y_int]
        if cat_up == "PERMANENTE":
            bk["permanente"].add(pid)
        elif cat_up == "COLABORADOR":
            bk["colaborador"].add(pid)
        elif cat_up == "VISITANTE":
            bk["visitante"].add(pid)

wb5b.close()

# ── 5b.3 Agregação ──────────────────────────────────────────────────────────
for iees in IEES:
    key = iees.lower()

    # Discentes — ano mais recente
    disc_yrs  = sorted(disc_raw.get(iees, {}).keys(), reverse=True)
    disc_y    = disc_yrs[0] if disc_yrs else None
    disc_data = disc_raw.get(iees, {}).get(disc_y, {}) if disc_y else {}

    # Docentes — ano mais recente
    doc_yrs  = sorted(doc_raw.get(iees, {}).keys(), reverse=True)
    doc_y    = doc_yrs[0] if doc_yrs else None
    doc_data = doc_raw.get(iees, {}).get(doc_y, {}) if doc_y else {}

    if disc_y:
        mat_m         = len(disc_data.get("mat_m", set()))
        mat_d         = len(disc_data.get("mat_d", set()))
        tit_m         = len(disc_data.get("tit_m", set()))
        tit_d         = len(disc_data.get("tit_d", set()))
        progs_conc    = disc_data.get("progs_conc", {})
        total_progs_d = len(progs_conc)
        progs_exc     = sum(1 for c in progs_conc.values() if c >= 6)
        pct_exc       = round(progs_exc / total_progs_d * 100, 2) if total_progs_d > 0 else None
        results[key].update(
            discMestrado=mat_m,
            discDoutorado=mat_d,
            tituladosMestrado=tit_m,
            tituladosDoutorado=tit_d,
            pctExcelencia=pct_exc,
        )
        src_disc = f"Base CAPES- Pós-Graduação - Brasil.xlsx / Base_Discentes / AN_BASE={disc_y}"
        sources[key].update(
            discMestrado       =src_disc + " / set(ID_PESSOA) DS_GRAU_ACADEMICO_DISCENTE=MESTRADO + NM_SITUACAO_DISCENTE=MATRICULADO",
            discDoutorado      =src_disc + " / set(ID_PESSOA) DS_GRAU_ACADEMICO_DISCENTE=DOUTORADO + NM_SITUACAO_DISCENTE=MATRICULADO",
            tituladosMestrado  =src_disc + " / set(ID_PESSOA) DS_GRAU_ACADEMICO_DISCENTE=MESTRADO + NM_SITUACAO_DISCENTE=TITULADO",
            tituladosDoutorado =src_disc + " / set(ID_PESSOA) DS_GRAU_ACADEMICO_DISCENTE=DOUTORADO + NM_SITUACAO_DISCENTE=TITULADO",
            pctExcelencia      =src_disc + " / NM_PROGRAMA_IES distintos com CD_CONCEITO_PROGRAMA>=6 / total NM_PROGRAMA_IES × 100",
        )

    if doc_y:
        doc_perm  = len(doc_data.get("permanente",  set()))
        doc_colab = len(doc_data.get("colaborador", set()))
        doc_vis   = len(doc_data.get("visitante",   set()))
        results[key].update(
            docPermanentes=doc_perm,
            docColaboradores=doc_colab,
            docVisitantes=doc_vis,
        )
        src_doc = f"Base CAPES- Pós-Graduação - Brasil.xlsx / Base_Docentes / AN_BASE={doc_y}"
        sources[key].update(
            docPermanentes   =src_doc + " / set(ID_PESSOA) DS_CATEGORIA_DOCENTE=PERMANENTE",
            docColaboradores =src_doc + " / set(ID_PESSOA) DS_CATEGORIA_DOCENTE=COLABORADOR",
            docVisitantes    =src_doc + " / set(ID_PESSOA) DS_CATEGORIA_DOCENTE=VISITANTE",
        )

    if disc_y and doc_y:
        doc_perm_n  = len(doc_data.get("permanente", set()))
        total_disc  = len(disc_data.get("mat_m", set())) + len(disc_data.get("mat_d", set()))
        razao       = round(total_disc / doc_perm_n, 2) if doc_perm_n > 0 else None
        results[key]["razaoDocenteDiscente"] = razao
        sources[key]["razaoDocenteDiscente"] = (
            f"(discMestrado + discDoutorado) / docPermanentes"
            f" — disc AN_BASE={disc_y}, doc AN_BASE={doc_y}"
        )


# ── 6. Orçamento — budget, execution, liquidation, personnel ─────────────────
# Fonte: Relatório da Despesa 8050 (2024 - 2026).xlsx / 2024-2026
#
# Colunas usadas (índices verificados em 2025-06):
#   [0]  Exercício       → ano
#   [47] Liquidado       → valor liquidado por linha de despesa (R$)
#   [49] Co_IES          → código inteiro da IES (mapeado via CO_IES_MAP)
#   [50] Taxa de Execução Orçamentária (Empenho)  → decimal (ex: 0.947)
#   [51] Taxa de Liquidação                        → decimal
#   [55] Participação de Pessoal e Encargos no Total da Despesa → decimal
#
# Regras:
#   budget      = soma(Liquidado) por IES/ano ÷ 1_000_000   (R$ milhões)
#   execution   = primeiro valor não-nulo da taxa por (IES, ano) selecionado
#   liquidation = idem
#   personnel   = idem; fallback 0.7034 (70,34%) se ausente
#   Ano preferido: 2024. Se não houver, usa o mais recente disponível.

_DESPESA_FILE = "Relatório da Despesa 8050 (2024 - 2026).xlsx"
_DESPESA_SHEET = "2024-2026"
_PERSONNEL_FALLBACK = 0.7034  # 70,34% — usado quando a IES não tem dado na coluna

wb = openpyxl.load_workbook(DATA_DIR / _DESPESA_FILE, read_only=True, data_only=True)
ws = wb[_DESPESA_SHEET]
_hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
_col = {h: i for i, h in enumerate(_hdr) if h is not None}

# Localiza cada coluna: índice fixo (verificado) com fallback por substring do nome,
# para sobreviver a eventuais realinhamentos do arquivo.
def _fcol(mapping, *keywords, default=None):
    """Retorna o índice da primeira coluna cujo nome contém todos os keywords."""
    for h, i in mapping.items():
        if h and all(k.lower() in str(h).lower() for k in keywords):
            return i
    return default

_co_col   = _fcol(_col, "Co_IES",    default=49)
_yr_col   = _fcol(_col, "Exerc",     default=0)
_liq_col  = _fcol(_col, "Liquidado", default=47)  # "Liquidado" ≠ "Liquidação"
_exec_col = _fcol(_col, "Execu", "Or",    default=50)
_liqr_col = _fcol(_col, "Taxa de Liquid", default=51)
_pes_col  = _fcol(_col, "Pessoal", "Particip", default=55)

# despesa_liq[(sigla, ano)]   = soma acumulada de Liquidado (R$)
# despesa_rates[(sigla, ano)] = {"exec", "liq", "pes"} — primeiro não-nulo encontrado
despesa_liq   = {}
despesa_rates = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    # Identifica a IES pelo código inteiro em Co_IES
    try:
        co_int = int(row[_co_col])
    except (TypeError, ValueError):
        continue
    sigla = CO_IES_MAP.get(co_int)
    if sigla not in IEES_PR:
        continue  # processa apenas as 7 IES paranaenses

    try:
        ano = int(row[_yr_col])
    except (TypeError, ValueError):
        continue

    k = (sigla, ano)

    # Acumula Liquidado (pode haver múltiplas linhas de despesa por IES/ano)
    try:
        despesa_liq[k] = despesa_liq.get(k, 0.0) + float(row[_liq_col])
    except (TypeError, ValueError):
        pass

    # Para as taxas, guarda o primeiro valor não-nulo por (IES, ano)
    if k not in despesa_rates:
        despesa_rates[k] = {"exec": None, "liq": None, "pes": None}
    r = despesa_rates[k]
    if r["exec"] is None:
        try:
            r["exec"] = float(row[_exec_col])
        except (TypeError, ValueError):
            pass
    if r["liq"] is None:
        try:
            r["liq"] = float(row[_liqr_col])
        except (TypeError, ValueError):
            pass
    if r["pes"] is None:
        try:
            r["pes"] = float(row[_pes_col])
        except (TypeError, ValueError):
            pass

wb.close()

_DESPESA_SRC = f"{_DESPESA_FILE} / {_DESPESA_SHEET}"

for iees in IEES_PR:
    key = iees.lower()
    budget = exec_r = liq_r = pes_r = year_used = None

    # Ordena os anos disponíveis: 2024 tem prioridade, depois decrescente
    anos_ies = sorted({ano for (s, ano) in despesa_liq if s == iees}, reverse=True)
    if 2024 in anos_ies:
        anos_ies = [2024] + [a for a in anos_ies if a != 2024]

    for ano in anos_ies:
        liq_total = despesa_liq.get((iees, ano), 0.0)
        if liq_total > 0:
            budget    = round(liq_total / 1_000_000, 2)
            d         = despesa_rates.get((iees, ano), {})
            exec_r    = d.get("exec")
            liq_r     = d.get("liq")
            pes_r     = d.get("pes")
            year_used = ano
            break  # ano preferido encontrado

    # personnel: usa valor da coluna (per-IES) ou fallback global 70,34%
    pes_final = pes_r if pes_r is not None else _PERSONNEL_FALLBACK

    results[key].update(
        budget=budget,
        execution=safe_pct(exec_r),
        liquidation=safe_pct(liq_r),
        personnel=round(pes_final * 100, 2),
    )

    _pes_src = (
        f"{_DESPESA_SRC} / Participação de Pessoal e Encargos / ano={year_used or 'N/D'} (per-IES)"
        if pes_r is not None
        else f"{_DESPESA_SRC} / fallback global {_PERSONNEL_FALLBACK * 100:.2f}%"
    )
    sources[key].update(
        budget=f"{_DESPESA_SRC} / sum(Liquidado) / Co_IES={iees} / ano={year_used or 'N/D'} (R$ milhões)",
        execution=f"{_DESPESA_SRC} / Taxa de Execução Orçamentária (Empenho) / ano={year_used or 'N/D'}",
        liquidation=f"{_DESPESA_SRC} / Taxa de Liquidação / ano={year_used or 'N/D'}",
        personnel=_pes_src,
    )


# ── 7. Suplementação ─────────────────────────────────────────────────────────
# Fonte: Dados de Suplementação das Universidades - Paraná.xlsx / matriz_cluster
# Coluna: col[3] — "% Suplementação" (proporção histórica de suplementações)

wb = openpyxl.load_workbook(DATA_DIR / "Dados de Suplementação das Universidades - Paraná.xlsx", read_only=True, data_only=True)
ws = wb["matriz_cluster"]
next(ws.iter_rows(min_row=1, max_row=1))  # skip header
for row in ws.iter_rows(min_row=2, values_only=True):
    iees = row[0] if row else None
    if iees not in IEES_PR:  # base Paraná
        continue
    supl = row[3] if len(row) > 3 else None
    key  = iees.lower()
    results[key]["supplementation"] = safe_pct(supl)
    sources[key]["supplementation"] = (
        "Dados de Suplementação das Universidades - Paraná.xlsx"
        " / matriz_cluster / col[3] % Suplementação (histórico)"
    )
wb.close()


# ── 8. CBO2/RAIS — taxa de inserção e salário médio ──────────────────────────
# Fonte: CBO2 _ RAIS 2023 e 2024 - Paraná.xlsx / Análise Quantitativa (BI e Cons
# Colunas (0-based):
#   0=IES, 1=egressos2020, 4=enc_PR_2023(CBO2), 5=sal_2023,
#          6=egressos2021, 9=enc_PR_2024(CBO2), 10=sal_2024,
#          11=egressos2022, 15=enc_PR_2025(CBO2), 16=sal_2025
# Transformação: employment = enc_PR ÷ egressos × 100
# Preferência: coorte 2022/RAIS2025 → 2021/RAIS2024 → 2020/RAIS2023

wb = openpyxl.load_workbook(DATA_DIR / "CBO2 _ RAIS 2023 e 2024 - Paraná.xlsx", read_only=True, data_only=True)
ws = wb["Análise Quantitativa (BI e Cons"]
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

for row in ws.iter_rows(min_row=2, max_row=15, values_only=True):
    if not row or row[0] is None:
        continue
    iees = str(row[0]).strip().upper()
    if iees not in IEES_PR:  # base Paraná
        continue
    key = iees.lower()

    eg_2022    = row[11]
    enc_pr_25  = row[15]
    sal_2025   = row[16]
    eg_2021    = row[6]
    enc_pr_24  = row[9]
    sal_2024   = row[10]
    eg_2020    = row[1]
    enc_pr_23  = row[4]
    sal_2023   = row[5]

    emp_rate = None
    sal_src  = None
    sal_val  = None

    def _try_coorte(eg, enc, sal, src):
        try:
            if eg and enc and isinstance(enc, (int, float)) and float(eg) > 0:
                return float(enc) / float(eg), src, sal
        except Exception:
            pass
        return None, None, None

    emp_rate, sal_src, sal_val = _try_coorte(eg_2022, enc_pr_25, sal_2025, "coorte 2022 / RAIS 2025")
    if emp_rate is None:
        emp_rate, sal_src, sal_val = _try_coorte(eg_2021, enc_pr_24, sal_2024, "coorte 2021 / RAIS 2024")
    if emp_rate is None:
        emp_rate, sal_src, sal_val = _try_coorte(eg_2020, enc_pr_23, sal_2023, "coorte 2020 / RAIS 2023")

    salary = None
    try:
        if isinstance(sal_val, (int, float)):
            salary = safe_float(sal_val, 0)
    except Exception:
        pass

    results[key]["employment"] = safe_pct(emp_rate)
    results[key]["salary"]     = salary
    sources[key]["employment"] = (
        f"CBO2 _ RAIS 2023 e 2024 - Paraná.xlsx"
        f" / Análise Quantitativa / enc_PR ÷ egressos × 100 / {sal_src or 'N/D'}"
    )
    sources[key]["salary"] = (
        f"CBO2 _ RAIS 2023 e 2024 - Paraná.xlsx"
        f" / Análise Quantitativa / média salarial egressos PR+CBO2 / {sal_src or 'N/D'}"
    )
wb.close()


# ── 9. Egressos — insertionRatePR ─────────────────────────────────────────────
# Fonte: Base Egressos - Paraná.xlsx / Base_Egressos_PR
# Colunas (0-based): 0=CO_IES, 1=IES, 2=Coorte, 3=Ano_RAIS,
#   [13] Taxa de inserção de egressos (Sul/BR, inclui SC+RS)  → insertionRatePR
# Difere de `employment` (col[15] PR-only + filtro CBO2); insertionRatePR é a
# taxa de reinserção formal mais ampla — egressos encontrados em qualquer estado do Sul.
# Prefere coorte 2021/RAIS 2024; fallback coorte 2020/RAIS 2023.
# Exclui coorte 2022 (RAIS 2025 ainda com erros #VALUE! no arquivo).

wb = openpyxl.load_workbook(DATA_DIR / "Base Egressos - Paraná.xlsx", read_only=True, data_only=True)
ws = wb["Base_Egressos_PR"]
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

egr_data = {}  # {iees: {(coorte, ano_rais): taxa_pr}}
for row in ws.iter_rows(min_row=2, values_only=True):
    co = row[0]
    try:
        co_int = int(co)
    except (TypeError, ValueError):
        continue
    iees = CO_IES_MAP.get(co_int)
    if iees not in IEES_PR:
        continue
    try:
        coorte   = int(row[2])
        ano_rais = int(row[3])
    except (TypeError, ValueError):
        continue
    if coorte == 2022:  # dados incompletos no arquivo
        continue
    taxa_sul = row[13]
    if not isinstance(taxa_sul, (int, float)):
        continue
    if iees not in egr_data:
        egr_data[iees] = {}
    egr_data[iees][(coorte, ano_rais)] = float(taxa_sul)
wb.close()

for iees in IEES_PR:
    key = iees.lower()
    if iees not in egr_data:
        continue
    pares = sorted(egr_data[iees].keys(), reverse=True)
    if not pares:
        continue
    best = pares[0]
    taxa = egr_data[iees][best]
    results[key]["insertionRatePR"] = round(taxa * 100, 2)
    sources[key]["insertionRatePR"] = (
        f"Base Egressos - Paraná.xlsx / Base_Egressos_PR"
        f" / Taxa de inserção de egressos (Sul — SC+PR+RS)"
        f" / coorte={best[0]} RAIS={best[1]}"
    )


# ── 11. Base RAIS — egressosMunicipios ────────────────────────────────────────
# Fonte: Base RAIS - 2023 e 2024 - Paraná.xlsx / Base_RAIS_2023_2024
# Colunas: [1] ANO_EGRESSO, [2] ANO_RAIS, [3] IEES, [12] MUNICIPIO_NOME
# egressosMunicipios = nº de municípios distintos onde egressos estão empregados
# Prefere o par (coorte, ano_rais) mais recente disponível.

wb = openpyxl.load_workbook(DATA_DIR / "Base RAIS - 2023 e 2024 - Paraná.xlsx", read_only=True, data_only=True)
ws = wb["Base_RAIS_2023_2024"]
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

rais_mun = {}  # {iees: {(coorte, ano_rais): set(municipios)}}
for row in ws.iter_rows(min_row=2, values_only=True):
    iees = str(row[3]).strip().upper() if row[3] else None
    if iees not in IEES_PR:
        continue
    try:
        ano_eg   = int(row[1])
        ano_rais = int(row[2])
    except (TypeError, ValueError):
        continue
    municipio = row[12]
    if not municipio or str(municipio).strip() == "":
        continue
    if iees not in rais_mun:
        rais_mun[iees] = {}
    pair = (ano_eg, ano_rais)
    if pair not in rais_mun[iees]:
        rais_mun[iees][pair] = set()
    rais_mun[iees][pair].add(str(municipio).strip())
wb.close()

for iees in IEES_PR:
    key = iees.lower()
    if iees not in rais_mun:
        continue
    pares = sorted(rais_mun[iees].keys(), reverse=True)
    if not pares:
        continue
    best = pares[0]
    results[key]["egressosMunicipios"] = len(rais_mun[iees][best])
    sources[key]["egressosMunicipios"] = (
        f"Base RAIS - 2023 e 2024 - Paraná.xlsx / Base_RAIS_2023_2024"
        f" / MUNICIPIO_NOME distintos / coorte={best[0]} RAIS={best[1]}"
    )


# ── 12. Estratificação — clusters V1-V8 e referência de quartis ───────────────
# Fonte: Estratificação_IES_Estaduais_BR.xlsx
# Sheet 1_Matriz de Estratificação: V1-V8 labels por IES (linha de dados a partir da linha 6)
# Sheet 0_Referência de Quartis: limiares e rótulos dos 4 quartis por variável
#
# Layout 0-indexed (linha de dados, sheet 1):
#   [2]=Sigla  [8]=V1  [11]=V2  [14]=V3  [17]=V4  [20]=V5  [22]=V6  [24]=V7  [26]=V8
#
# "Não disponível" → None  (V6-V8 só existem para IES-PR)

def _strat_label(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ("não disponível", "nao disponivel", "n/a", ""):
        return None
    return s

def _strat_float(v, digits=None):
    if v is None or v == "":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return round(n, digits) if digits is not None else n

wb = openpyxl.load_workbook(
    DATA_DIR / "Estratificação_IES_Estaduais_BR.xlsx", read_only=True, data_only=True
)

# 12a. Grupos por IES
ws_mat = wb["1_Matriz de Estratificação"]
clusters_raw = {}
for row in ws_mat.iter_rows(min_row=6, values_only=True):
    sigla = row[2] if len(row) > 2 else None
    if sigla not in IEES:
        continue
    clusters_raw[sigla] = {
        "v1": _strat_label(row[8]  if len(row) > 8  else None),
        "v2": _strat_label(row[11] if len(row) > 11 else None),
        "v3": _strat_label(row[14] if len(row) > 14 else None),
        "v4": _strat_label(row[17] if len(row) > 17 else None),
        "v5": _strat_label(row[20] if len(row) > 20 else None),
        "v6": _strat_label(row[22] if len(row) > 22 else None),
        "v7": _strat_label(row[24] if len(row) > 24 else None),
        "v8": _strat_label(row[26] if len(row) > 26 else None),
    }

# 12b. Referência de quartis (linha 6 em diante, até linha sem dados)
ws_ref = wb["0_Referência de Quartis"]
quartis_ref = []
for row in ws_ref.iter_rows(min_row=6, values_only=True):
    if not row[0]:
        continue
    # Ignora linhas de notas metodológicas (não são variáveis de quartil)
    if str(row[0]).startswith("NOTA") or row[5] is None:
        continue
    quartis_ref.append({
        "variable":   str(row[0]).strip(),
        "indicator":  str(row[1]).strip() if row[1] else None,
        "q1_limiar":  str(row[2]).strip() if row[2] else None,
        "q2_limiar":  str(row[3]).strip() if row[3] else None,
        "q3_limiar":  str(row[4]).strip() if row[4] else None,
        "label_q1":   str(row[5]).strip() if row[5] else None,
        "label_q2":   str(row[6]).strip() if row[6] else None,
        "label_q3":   str(row[7]).strip() if row[7] else None,
        "label_q4":   str(row[8]).strip() if row[8] else None,
    })

# 12c. Valores territoriais oficiais (PR) para V7 e V8.
# As abas 10/11 trazem valores ponderados por território atendido; estes campos
# substituem qualquer valor antigo baseado em município-sede.
for row in wb["9_Renda Território PR"].iter_rows(min_row=6, values_only=True):
    sigla = str(row[2]).strip() if len(row) > 2 and row[2] else None
    if sigla not in IEES_PR:
        continue
    key = sigla.lower()
    renda = _strat_float(row[5] if len(row) > 5 else None, 2)
    faixa = _strat_label(row[6] if len(row) > 6 else None)
    if renda is not None:
        results[key]["territoryIncome"] = renda
        sources[key]["territoryIncome"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 9_Renda Território PR"
            " / Renda Per Capita Ponderada (R$)"
        )
    if faixa:
        results[key]["v7_label"] = faixa
        sources[key]["v7_label"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 9_Renda Território PR"
            " / Faixa de Renda Territorial"
        )
        if sigla in clusters_raw:
            clusters_raw[sigla]["v7"] = faixa

for row in wb["10_IDH Território PR"].iter_rows(min_row=6, values_only=True):
    sigla = str(row[2]).strip() if len(row) > 2 and row[2] else None
    if sigla not in IEES_PR:
        continue
    key = sigla.lower()
    idh = _strat_float(row[5] if len(row) > 5 else None, 4)
    faixa = _strat_label(row[6] if len(row) > 6 else None)
    if idh is not None:
        results[key]["idhmRegional"] = idh
        sources[key]["idhmRegional"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 10_IDH Território PR"
            " / IDH Municipal Ponderado (0–1)"
        )
    if faixa:
        results[key]["v8_label"] = faixa
        sources[key]["v8_label"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 10_IDH Território PR"
            " / Faixa de Contexto Socioeconômico"
        )
        if sigla in clusters_raw:
            clusters_raw[sigla]["v8"] = faixa

# 12d. V9 — Área de Atuação Predominante (Grande Área CINE). Usada só como
# filtro global de UNIVERSO de IES (Round 2a) — NÃO é variável de groupBy
# V1-V8 e NÃO entra em clusters_raw (mantido intocado, ver linha 1380).
# Reaproveita ws_mat (mesmo workbook já aberto em 12a) numa segunda
# passada — sem reabrir o arquivo.
for row in ws_mat.iter_rows(min_row=6, values_only=True):
    sigla = row[2] if len(row) > 2 else None
    if sigla not in IEES:
        continue
    key = sigla.lower()
    area = _strat_label(row[27] if len(row) > 27 else None)
    pct = _strat_float(row[28] if len(row) > 28 else None, 2)
    herf = _strat_float(row[29] if len(row) > 29 else None, 3)
    if area is not None:
        results[key]["areaCineGrande"] = area
        sources[key]["areaCineGrande"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 1_Matriz de Estratificação"
            " / V9 – Área de Atuação Predominante (col 27)"
        )
    if pct is not None:
        results[key]["areaCinePct"] = pct
        sources[key]["areaCinePct"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 1_Matriz de Estratificação"
            " / V9 – % Matrículas na Área (col 28)"
        )
    if herf is not None:
        results[key]["areaCineHerfindahl"] = herf
        sources[key]["areaCineHerfindahl"] = (
            "Estratificação_IES_Estaduais_BR.xlsx / 1_Matriz de Estratificação"
            " / V9 – Índice de Concentração / Herfindahl (col 29)"
        )

wb.close()


# ── Seção 9 — Relatório Despesa 8050 ─────────────────────────────────────────
# Extrai campos financeiros (soma por IES/ano) e indicadores pré-calculados
# (primeira linha de cada IES/ano) da aba "2024-2026".
#
# Identificação: col[6] = Unidade Orçamentária (UO) — distinto de col[49]=Co_IES.
# Cobre anos 2024, 2025 e 2026.
#
# Campos financeiros (cols 33,37,45,47,48): somados e convertidos para R$ milhões.
# Taxas/participações (cols 50-64): captadas da primeira linha de cada (IES, ano)
#   e convertidas de decimal (0-1) para % (× 100).
#
# Saída:
#   - results[key] ← campos 2024 mesclados (flat, para o dashboard)
#   - d8050_by_year ← {sigla: {str(ano): {fields}}} para 2024/2025/2026

# UO → sigla (aba De-para do arquivo)
_UO_IES_MAP = {
    4530: "UEL",
    4531: "UEPG",
    4532: "UEM",
    4533: "UNICENTRO",
    4534: "UNIOESTE",
    4546: "UNESPAR",
    4548: "UENP",
}

# Colunas financeiras: acumular soma por (sigla, ano)
_D8050_FIN_COLS = {
    "dotacao_inicial":      33,
    "orcamento_atualizado": 37,
    "empenhado":            45,
    "liquidado":            47,
    "pago":                 48,
}

# Colunas de taxas/participações: guardar primeira linha por (sigla, ano)
# Valor original: decimal 0-1 → gravar como % (× 100)
_D8050_RATE_COLS = {
    "tx_execucao_empenho":    50,
    "tx_liquidacao":          51,
    "tx_pagamento_liq":       52,
    "grau_contingenciamento": 53,
    "var_dotacao_loa":        54,
    "part_pessoal":           55,
    "part_outras_correntes":  56,
    "part_capital":           57,
    "part_recursos_livres":   58,
    "part_fonte_500":         59,
    "part_fonte_501":         60,
    "part_demais_vincul":     61,
    "part_convenios_uniao":   62,
    "part_convenios_privados":63,
    "part_emendas_federais":  64,
}

_D8050_ANOS = {2024, 2025, 2026}

d8050_fin   = {}   # {(sigla, ano): {field: float_acumulado}}
d8050_rates = {}   # {(sigla, ano): {field: float|None}}  — primeira linha

wb9 = openpyxl.load_workbook(DATA_DIR / _DESPESA_FILE, read_only=True, data_only=True)
ws9 = wb9[_DESPESA_SHEET]
next(ws9.iter_rows(min_row=1, max_row=1))  # pula cabeçalho

for row in ws9.iter_rows(min_row=2, values_only=True):
    # Identificar IES pela Unidade Orçamentária (col 6)
    try:
        uo = int(row[6])
    except (TypeError, ValueError):
        continue
    sigla = _UO_IES_MAP.get(uo)
    if sigla is None:
        continue

    try:
        ano = int(row[0])
    except (TypeError, ValueError):
        continue
    if ano not in _D8050_ANOS:
        continue

    k = (sigla, ano)

    # Acumula campos financeiros
    if k not in d8050_fin:
        d8050_fin[k] = {f: 0.0 for f in _D8050_FIN_COLS}
    for field, col in _D8050_FIN_COLS.items():
        try:
            v = float(row[col]) if row[col] is not None else 0.0
            d8050_fin[k][field] += v
        except (TypeError, ValueError):
            pass

    # Taxas — apenas a primeira linha encontrada por (sigla, ano)
    if k not in d8050_rates:
        d8050_rates[k] = {f: None for f in _D8050_RATE_COLS}
    for field, col in _D8050_RATE_COLS.items():
        if d8050_rates[k][field] is None:
            try:
                v = float(row[col])
                d8050_rates[k][field] = v
            except (TypeError, ValueError):
                pass

wb9.close()

# Montar estrutura multi-year: {sigla: {str(ano): {fields}}}
d8050_by_year = {}
_D8050_SRC = f"{_DESPESA_FILE} / {_DESPESA_SHEET} / col[6]=UO"

for sigla in IEES_PR:
    d8050_by_year[sigla] = {}
    for ano in sorted(_D8050_ANOS):
        k = (sigla, ano)
        if k not in d8050_fin and k not in d8050_rates:
            continue
        yr_data = {}

        # Campos financeiros → R$ milhões (2 casas)
        if k in d8050_fin:
            for field in _D8050_FIN_COLS:
                raw = d8050_fin[k].get(field, 0.0)
                yr_data[field] = round(raw / 1_000_000, 2) if raw else None

        # Taxas → % (2 casas)
        if k in d8050_rates:
            for field in _D8050_RATE_COLS:
                raw = d8050_rates[k].get(field)
                yr_data[field] = round(raw * 100, 2) if raw is not None else None

        d8050_by_year[sigla][str(ano)] = yr_data

# Mesclar campos de 2024 no results flat (para o dashboard)
for sigla in IEES_PR:
    key = sigla.lower()
    fields_2024 = d8050_by_year.get(sigla, {}).get("2024", {})
    if fields_2024:
        results[key].update(fields_2024)
        src = f"{_D8050_SRC} / ano=2024"
        for field in fields_2024:
            sources[key][field] = src

# ── Aliases ind81–ind87 (Eficiência e Estrutura Orçamentária) ─────────────────
# Mapeados dos campos D8050 já extraídos. Escopo: 7 IES-PR, ano 2024.
_IND8X_MAP = {
    "ind81": "tx_execucao_empenho",   # Execução (Empenho)
    "ind82": "tx_liquidacao",          # Liquidação
    "ind83": "tx_pagamento_liq",       # Pagamento/Liquidado
    "ind84": "grau_contingenciamento", # Contingenciamento ↓ melhor
    "ind85": "var_dotacao_loa",        # Variação Dotação vs LOA inicial
    "ind86": "part_pessoal",           # Participação Pessoal e Encargos
    "ind87": "part_outras_correntes",  # Participação Outras Despesas Correntes
}
for sigla in IEES_PR:
    key = sigla.lower()
    for ind, campo in _IND8X_MAP.items():
        val = results[key].get(campo)
        if val is not None:
            results[key][ind] = val
            sources[key][ind] = sources[key].get(campo, "")


# ── Seção 10 — Estratificação V6 (Dinâmica Orçamentária PR) ──────────────────
# Lê a aba '8_Dinâmica Orçamentária PR' do arquivo de estratificação e extrai
# o índice composto e a faixa de perfil orçamentário por IES-PR.
# Substitui clusters_raw[sigla]["v6"] com o rótulo oficial.
#
# Estrutura da aba (cabeçalho na linha 5, dados a partir da 6):
#   col [2]  = Sigla
#   col [13] = Índice Composto (0–1)
#   col [14] = Faixa de Perfil Orçamentário
#
_wb_v6 = openpyxl.load_workbook(
    DATA_DIR / "Estratificação_IES_Estaduais_BR.xlsx", read_only=True, data_only=True
)
_ws_v6 = _wb_v6["8_Dinâmica Orçamentária PR"]

for _row in _ws_v6.iter_rows(min_row=6, values_only=True):
    _sigla = str(_row[2]).strip() if len(_row) > 2 and _row[2] else None
    if _sigla not in IEES_PR:
        continue

    try:
        _v6_indice = round(float(_row[13]), 4) if len(_row) > 13 and _row[13] is not None else None
    except (TypeError, ValueError):
        _v6_indice = None
    _v6_perfil = str(_row[14]).strip() if len(_row) > 14 and _row[14] else None
    _src = "Estratificação_IES_Estaduais_BR.xlsx / 8_Dinâmica Orçamentária PR"

    _key = _sigla.lower()
    if _v6_indice is not None:
        results[_key]["v6_indice"] = _v6_indice
        sources[_key]["v6_indice"] = _src
    if _v6_perfil:
        results[_key]["v6_perfil"] = _v6_perfil
        sources[_key]["v6_perfil"] = _src
        if _sigla in clusters_raw:
            clusters_raw[_sigla]["v6"] = _v6_perfil

_wb_v6.close()

_SEP10 = "─" * 60
print("", file=sys.stderr)
print(_SEP10, file=sys.stderr)
print("Seção 10 — V6 Dinâmica Orçamentária PR | Validação", file=sys.stderr)
print(_SEP10, file=sys.stderr)
print(f"{'IES':<12} {'v6_indice':>12}  v6_perfil", file=sys.stderr)
print(_SEP10, file=sys.stderr)
for _s in IEES_PR:
    _k   = _s.lower()
    _idx  = results[_k].get("v6_indice")
    _perf = results[_k].get("v6_perfil", "N/D")
    _idx_str = f"{_idx:.4f}" if _idx is not None else "N/D"
    print(f"{_s:<12} {_idx_str:>12}  {_perf}", file=sys.stderr)
print(_SEP10, file=sys.stderr)


# ── Seção 13 — Composição de Fontes de Despesa ───────────────────────────────
# Derivado dos campos já extraídos em d8050_by_year para 2024.
# Calcula pct_no_grupo e valor (R$ mi) para cada fonte dentro dos grupos 50 e 70.

_FONTES_NOMES = {
    "500": "Recursos do Tesouro",
    "501": "Arrecadação Própria",
    "700": "Convênios da União",
    "703": "Conv. Entidades Privadas",
    "706": "Emendas Federais",
}

def _composicao_fontes(yr_data):
    g50 = yr_data.get("part_recursos_livres")
    g70 = yr_data.get("part_demais_vincul")
    orc = yr_data.get("orcamento_atualizado")  # R$ milhões

    def _fonte(code, pct_orc, g_pct):
        if pct_orc is None:
            return None
        return {
            "nome": _FONTES_NOMES.get(code, code),
            "pct_no_orcamento": round(pct_orc, 2),
            "pct_no_grupo": round(pct_orc / g_pct * 100, 2) if g_pct else None,
            "valor": round(pct_orc / 100 * orc, 3) if orc else None,
        }

    def _grupo(nome, total_pct, fontes_raw):
        fontes = {k: v for k, v in fontes_raw.items() if v is not None}
        return {"nome": nome, "total_pct": total_pct, "fontes": fontes}

    return {
        "grupo50": _grupo(
            "Recursos Livres (não vinculados)", g50,
            {
                "500": _fonte("500", yr_data.get("part_fonte_500"), g50),
                "501": _fonte("501", yr_data.get("part_fonte_501"), g50),
            },
        ),
        "grupo70": _grupo(
            "Demais Vinculações Decorrentes", g70,
            {
                "700": _fonte("700", yr_data.get("part_convenios_uniao"), g70),
                "703": _fonte("703", yr_data.get("part_convenios_privados"), g70),
                "706": _fonte("706", yr_data.get("part_emendas_federais"), g70),
            },
        ),
    }

composicaoFontes = {}
for _sigla in IEES_PR:
    _yr = d8050_by_year.get(_sigla, {}).get("2024", {})
    if _yr:
        composicaoFontes[_sigla] = _composicao_fontes(_yr)


# ── Seção 13b — IND 80 (Dispersão Territorial dos Egressos) ──────────────────
# Fonte: Base RAIS - 2023 e 2024 - Paraná.xlsx / Base_RAIS_2023_2024
# Coluna [22]: Índice de dispersão territorial dos egressos por curso (pré-calculado)
# ind80 = média do índice por IES, par (coorte, ano_rais) mais recente
# Valores típicos: 0.004–0.015 (escala original do arquivo)
# Escopo: 7 IES-PR

wb_rais80 = openpyxl.load_workbook(DATA_DIR / "Base RAIS - 2023 e 2024 - Paraná.xlsx", read_only=True, data_only=True)
ws_rais80 = wb_rais80["Base_RAIS_2023_2024"]
next(ws_rais80.iter_rows(min_row=1, max_row=1))  # pula header

_rais80_raw = {}  # {iees: {(coorte, ano_rais): [disp_vals]}}
for _row in ws_rais80.iter_rows(min_row=2, values_only=True):
    _iees = str(_row[3]).strip().upper() if _row[3] else None
    if _iees not in IEES_PR:
        continue
    try:
        _ano_eg = int(_row[1]); _ano_rais = int(_row[2])
    except (TypeError, ValueError):
        continue
    _disp = _row[22]
    if _disp is None:
        continue
    try:
        _disp = float(_disp)
    except (TypeError, ValueError):
        continue
    if _iees not in _rais80_raw:
        _rais80_raw[_iees] = {}
    _par = (_ano_eg, _ano_rais)
    if _par not in _rais80_raw[_iees]:
        _rais80_raw[_iees][_par] = []
    _rais80_raw[_iees][_par].append(_disp)
wb_rais80.close()

_SRC_RAIS80 = "Base RAIS - 2023 e 2024 - Paraná.xlsx / Base_RAIS_2023_2024 / [22] Índice dispersão territorial / média por IES (par mais recente)"
for _iees in IEES_PR:
    _key = _iees.lower()
    if _iees not in _rais80_raw:
        results[_key]["ind80"] = None
        sources[_key]["ind80"] = _SRC_RAIS80 + " / base insuficiente"
        continue
    _pares = sorted(_rais80_raw[_iees].keys(), reverse=True)
    if not _pares:
        results[_key]["ind80"] = None
        continue
    _best = _pares[0]
    _vals = _rais80_raw[_iees][_best]
    _media = sum(_vals) / len(_vals)
    # Multiplica por 100 para apresentar em % (0.0083 → 0.83%)
    results[_key]["ind80"] = round(_media * 100, 4)
    sources[_key]["ind80"] = _SRC_RAIS80 + f" / coorte={_best[0]} RAIS={_best[1]} / n_cursos={len(_vals)}"


# ── Seção 14 — IND 88–95 (Estrutura e Capacidade de Investimento) ────────────
# ind88: Razão Correntes/Capital — Cat 3 vs Cat 4
# ind89: Recursos Livres → alias de part_recursos_livres
# ind90: Recursos Próprios → alias de part_fonte_501
# ind91: Transferências → alias de part_demais_vincul
# ind92: Obras (Elemento 51, Liquidado / Orçamento Atualizado × 100)
# ind93: Equipamentos (Elemento 52, Liquidado / Orçamento Atualizado × 100)
# ind94: Variação Dotação → alias de var_dotacao_loa (= ind85)
# ind95: Execução sobre LOA Inicial (Liquidado / DotaçãoInicial × 100)
# Colunas (0-indexed): [0]=Ano [6]=UO [13]=Categoria [16]=Elemento
#   [33]=DotaçãoInicial [37]=OrcAtualizado [47]=Liquidado [49]=Co_IES

# ind89, ind90, ind91, ind94: aliases de campos já existentes em results
_IND9X_DIRECT = {
    "ind89": "part_recursos_livres",
    "ind90": "part_fonte_501",
    "ind91": "part_demais_vincul",
    "ind94": "var_dotacao_loa",
}
for _sig in IEES_PR:
    _k = _sig.lower()
    for _ind, _campo in _IND9X_DIRECT.items():
        _v = results[_k].get(_campo)
        if _v is not None:
            results[_k][_ind] = _v
            sources[_k][_ind] = sources[_k].get(_campo, "")

# ind88, ind92, ind93, ind95: lidos do arquivo D8050
_inv_data = {
    sig: {"orc_corr": 0.0, "orc_cap": 0.0,
          "liq_51": 0.0, "liq_52": 0.0,
          "liq_total": 0.0, "dot_total": 0.0}
    for sig in IEES_PR
}

wb_inv = openpyxl.load_workbook(DATA_DIR / _DESPESA_FILE, read_only=True, data_only=True)
ws_inv = wb_inv[_DESPESA_SHEET]
for _row in ws_inv.iter_rows(min_row=2, values_only=True):
    try:
        _uo = int(_row[6]); _ano = int(_row[0])
    except Exception:
        continue
    _sig = _UO_IES_MAP.get(_uo)
    if _sig is None or _ano != 2024:
        continue
    _d = _inv_data[_sig]
    _cat = str(_row[13]) if _row[13] is not None else ""
    _el  = str(_row[16]) if _row[16] is not None else ""
    def _f(v):
        try: return float(v) if v else 0.0
        except: return 0.0
    _orc = _f(_row[37]); _liq = _f(_row[47]); _dot = _f(_row[33])
    if _cat == "3": _d["orc_corr"] += _orc
    if _cat == "4": _d["orc_cap"]  += _orc
    if _el  == "51": _d["liq_51"]  += _liq
    if _el  == "52": _d["liq_52"]  += _liq
    _d["liq_total"] += _liq
    _d["dot_total"]  += _dot
wb_inv.close()

_SRC_INV = f"{_DESPESA_FILE} / {_DESPESA_SHEET} / ano=2024"
for _sig in IEES_PR:
    _k = _sig.lower()
    _d = _inv_data[_sig]
    _orc_total = (_d["orc_corr"] + _d["orc_cap"]) or 1.0
    # ind88: razão Correntes/Capital (não percentual — razão ex: 22.6)
    if _d["orc_cap"] > 0:
        results[_k]["ind88"] = round(_d["orc_corr"] / _d["orc_cap"], 2)
        sources[_k]["ind88"] = _SRC_INV + " / OrcCorrente÷OrcCapital (Cat3÷Cat4)"
    # ind92: Obras (Elemento 51)
    results[_k]["ind92"] = round(_d["liq_51"] / _orc_total * 100, 2)
    sources[_k]["ind92"] = _SRC_INV + " / Liquidado Elemento 51 ÷ OrcAtualizado × 100"
    # ind93: Equipamentos (Elemento 52)
    results[_k]["ind93"] = round(_d["liq_52"] / _orc_total * 100, 2)
    sources[_k]["ind93"] = _SRC_INV + " / Liquidado Elemento 52 ÷ OrcAtualizado × 100"
    # ind95: Execução sobre LOA Inicial
    if _d["dot_total"] > 0:
        results[_k]["ind95"] = round(_d["liq_total"] / _d["dot_total"] * 100, 2)
        sources[_k]["ind95"] = _SRC_INV + " / sum(Liquidado) ÷ sum(DotaçãoInicial) × 100"


# ── Seção 15 — Aliases ind81–87 nos anos 2025 e 2026 (série histórica) ───────
# d8050_by_year já tem tx_execucao_empenho etc. para 2024/2025/2026.
# Adiciona aliases ind81–87 nos yearData de cada ano para que o dashboard
# possa acessar pelo nome ind81 independente do ano selecionado.
for _sig in IEES_PR:
    for _yr in ["2024", "2025", "2026"]:
        _yr_data = d8050_by_year.get(_sig, {}).get(_yr, {})
        if not _yr_data:
            continue
        for _ind, _campo in _IND8X_MAP.items():
            _v = _yr_data.get(_campo)
            if _v is not None:
                _yr_data[_ind] = _v


# ── Seção 16 — Base SELO-PR: avaliação anual da execução orçamentária ────────
#
# O SELO Paraná (Sistema de Excelência em Liderança Orçamentária) é uma avaliação
# institucional da qualidade da execução orçamentária e financeira das universidades
# estaduais, conduzida pela Diretoria de Orçamento Estadual (DOE/SEFA-PR).
#
# A avaliação cobre três eixos temáticos:
#   Eixo I   — Eficiência na Execução Orçamentária (máx. 60 pts)
#              inds. 1.1 Empenho (12), 1.2 Liquidação (20),
#                    1.3 Empenho Liquidado (16), 1.4 Foco em Ações Finalísticas (12)
#   Eixo II  — Racionalidade na Gestão de Créditos Adicionais (máx. 20 pts)
#              inds. 2.1 Aderência à Programação (6), 2.2 Execução do Superávit (6),
#                    2.3 Priorização do Crédito do Exercício (8)
#   Eixo III — Passivos de Exercícios Anteriores (máx. 20 pts)
#              inds. 3.1 Inscrição em RAP (5), 3.2 Cancelamento de RAP (5),
#                    3.3 Pagamento de RAP (5), 3.4 Impacto de DEA (5)
#
# Nota final (0–100): pré-calculada pela DOE/SEFA e disponível na aba
# Nota_Final_Unidade. Cada indicador tem nota individual na aba Base_Indicadores.
#
# Fonte: Base SELO - Paraná.xlsx
# Abas lidas: Base_Indicadores, Nota_Final_Unidade

_SELO_FILE_CANDIDATES = [
    "SELO_PR_Base_Indicadores_Consolidada_uma_aba.xlsx",
    "SELO_PR_Base_Indicadores_Consolidada.xlsx",
    "Base SELO - Paraná.xlsx",
]

_SELO_INDICADORES = {
    "1.1": {"nome": "Empenho",                         "eixo": "I",   "maximo": 12, "polaridade": "maior"},
    "1.2": {"nome": "Liquidação",                       "eixo": "I",   "maximo": 20, "polaridade": "maior"},
    "1.3": {"nome": "Empenho Liquidado",                "eixo": "I",   "maximo": 16, "polaridade": "maior"},
    "1.4": {"nome": "Foco em Ações Finalísticas",       "eixo": "I",   "maximo": 12, "polaridade": "maior"},
    "2.1": {"nome": "Aderência à Programação Orç.",     "eixo": "II",  "maximo":  6, "polaridade": "maior"},
    "2.2": {"nome": "Execução do Superávit Concedido",  "eixo": "II",  "maximo":  6, "polaridade": "maior"},
    "2.3": {"nome": "Priorização do Crédito do Exerc.", "eixo": "II",  "maximo":  8, "polaridade": "maior"},
    "3.1": {"nome": "Inscrição em Restos a Pagar",      "eixo": "III", "maximo":  5, "polaridade": "menor"},
    "3.2": {"nome": "Cancelamento de Restos a Pagar",   "eixo": "III", "maximo":  5, "polaridade": "menor"},
    "3.3": {"nome": "Pagamento de Restos a Pagar",      "eixo": "III", "maximo":  5, "polaridade": "maior"},
    "3.4": {"nome": "Impacto de Despesas de Ex. Ant.",  "eixo": "III", "maximo":  5, "polaridade": "menor"},
}

_SELO_UNIDADES_ESPERADAS = set(IEES_PR)
# A base SELO-PR disponível está consolidada por indicador anual para 2025.
# Ela não traz desagregação bimestral; portanto a validação esperada é:
# 7 IEES × 1 ano × 11 indicadores = 77 linhas.
_SELO_ANOS_ESPERADOS = {2025}
_SELO_BIMESTRES_ESPERADOS = set()
_SELO_INDICADORES_ESPERADOS = set(_SELO_INDICADORES)


def _selo_norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    return "".join(ch if ch.isalnum() else "_" for ch in text).strip("_")


def _selo_find_file():
    for name in _SELO_FILE_CANDIDATES:
        path = DATA_DIR / name
        if path.exists():
            return path
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        norm = _selo_norm(path.name)
        if "selo" in norm and ("pr" in norm or "parana" in norm):
            return path
    return None


def _selo_headers(row):
    return [
        str(v).strip() if v not in (None, "") else f"Coluna {idx + 1}"
        for idx, v in enumerate(row or [])
    ]


def _selo_header_index(headers):
    index = {}
    for idx, header in enumerate(headers):
        key = _selo_norm(header)
        if key and key not in index:
            index[key] = idx
    return index


def _selo_col(header_idx, aliases, fallback=None):
    for alias in aliases:
        idx = header_idx.get(_selo_norm(alias))
        if idx is not None:
            return idx
    return fallback


def _selo_cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _selo_json_value(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def _selo_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _selo_year(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _selo_code(value):
    if value is None or value == "":
        return None
    text = str(value).strip().replace(",", ".")
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _selo_bimestre(value):
    text = _selo_text(value)
    if not text:
        return None
    norm = _selo_norm(text).upper()
    if norm.startswith("B") and norm[1:].isdigit():
        return f"B{int(norm[1:])}"
    digits = "".join(ch for ch in norm if ch.isdigit())
    return f"B{int(digits)}" if digits else text.upper()


def _selo_sigla(value):
    if value is None or value == "":
        return None
    try:
        uo = int(float(value))
        if uo in _UO_IES_MAP:
            return _UO_IES_MAP[uo]
    except (TypeError, ValueError):
        pass
    text = str(value).strip().upper()
    if text in IEES_PR:
        return text
    norm = _selo_norm(text)
    if "londrina" in norm:
        return "UEL"
    if "maringa" in norm:
        return "UEM"
    if "ponta_grossa" in norm:
        return "UEPG"
    if "centro_oeste" in norm or "unicentro" in norm:
        return "UNICENTRO"
    if "norte_do_parana" in norm or "uenp" in norm:
        return "UENP"
    if "oeste_do_parana" in norm or "unioeste" in norm:
        return "UNIOESTE"
    if "unespar" in norm or "estadual_do_parana" in norm:
        return "UNESPAR"
    return None


def _selo_record_from_row(row, headers, cols):
    unidade_raw = _selo_cell(row, cols["unidade"])
    ente_raw = _selo_cell(row, cols["ente"])
    sigla = _selo_sigla(unidade_raw) or _selo_sigla(ente_raw)
    return {
        "Ano": _selo_year(_selo_cell(row, cols["ano"])),
        "Ente Avaliado": _selo_text(ente_raw),
        "Unidade Orçamentária": sigla or _selo_text(unidade_raw),
        "Bimestre": _selo_bimestre(_selo_cell(row, cols["bimestre"])),
        "Peso do Bimestre": safe_float(_selo_cell(row, cols["peso_bimestre"]), 6),
        "Eixo": _selo_text(_selo_cell(row, cols["eixo"])),
        "Código Indicador": _selo_code(_selo_cell(row, cols["codigo"])),
        "Nome Indicador": _selo_text(_selo_cell(row, cols["nome_indicador"])),
        "Nota do Indicador no Bimestre": safe_float(_selo_cell(row, cols["nota_indicador"]), 6),
        "Percentual": safe_float(_selo_cell(row, cols["percentual"]), 6),
        "Nota do Bimestre": safe_float(_selo_cell(row, cols["nota_bimestre"]), 6),
        "Nota Final da Unidade": safe_float(_selo_cell(row, cols["nota_final"]), 6),
        "Nota Máxima do Indicador": safe_float(_selo_cell(row, cols["nota_maxima"]), 6),
        "Fonte": _selo_text(_selo_cell(row, cols["fonte"])),
        "Data de Extração": _selo_json_value(_selo_cell(row, cols["data_extracao"])),
        "_raw": {
            headers[idx]: _selo_json_value(row[idx] if idx < len(row) else None)
            for idx in range(len(headers))
        },
    }


def _selo_warn_missing(label, expected, found, alerts):
    missing = sorted(expected - found)
    if missing:
        alerts.append(f"{label} ausentes: {', '.join(map(str, missing))}")


caminho_selo = _selo_find_file()
if caminho_selo is None:
    raise FileNotFoundError("Base SELO-PR não encontrada em data/")
_SELO_FILE = caminho_selo.name

# Variáveis padronizadas da base SELO-PR (listas de registros do pipeline).
df_selo = []
df_selo_resumo = []
df_selo_bimestres = []
df_selo_indicadores = []

# {(sigla, ano): {"notaFinal": float, "indicadores": {"1.1": float, ...}}}
selo_data = {}
_selo_indicator_acc = {}
_selo_bimestre_acc = {}

wb_selo = openpyxl.load_workbook(caminho_selo, read_only=True, data_only=True)
_selo_sheet = "Base_Indicadores" if "Base_Indicadores" in wb_selo.sheetnames else wb_selo.sheetnames[0]
ws_ind = wb_selo[_selo_sheet]
_selo_header_row = next(ws_ind.iter_rows(min_row=1, max_row=1, values_only=True), ())
_selo_headers_list = _selo_headers(_selo_header_row)
_selo_hidx = _selo_header_index(_selo_headers_list)
_selo_cols = {
    "ano": _selo_col(_selo_hidx, ["Ano"], 0),
    "ente": _selo_col(_selo_hidx, ["Ente Avaliado", "Ente", "Instituição", "Instituicao"], None),
    "unidade": _selo_col(_selo_hidx, ["Unidade Orçamentária", "Unidade Orcamentaria", "IEES", "Sigla"], 2),
    "bimestre": _selo_col(_selo_hidx, ["Bimestre"], None),
    "peso_bimestre": _selo_col(_selo_hidx, ["Peso do Bimestre", "Peso Bimestre"], None),
    "eixo": _selo_col(_selo_hidx, ["Eixo"], None),
    "codigo": _selo_col(_selo_hidx, ["Código Indicador", "Codigo Indicador", "Cod Indicador"], 4),
    "nome_indicador": _selo_col(_selo_hidx, ["Nome Indicador", "Indicador"], None),
    "nota_indicador": _selo_col(
        _selo_hidx,
        ["Nota do Indicador no Bimestre", "Nota Indicador no Bimestre", "Nota do Indicador", "Nota Indicador"],
        6,
    ),
    "percentual": _selo_col(_selo_hidx, ["Percentual", "%"], None),
    "nota_bimestre": _selo_col(_selo_hidx, ["Nota do Bimestre", "Nota Bimestre"], None),
    "nota_final": _selo_col(_selo_hidx, ["Nota Final da Unidade", "Nota Final Unidade", "Nota Final"], None),
    "nota_maxima": _selo_col(_selo_hidx, ["Nota Máxima do Indicador", "Nota Maxima do Indicador", "Nota Máxima"], None),
    "fonte": _selo_col(_selo_hidx, ["Fonte"], None),
    "data_extracao": _selo_col(_selo_hidx, ["Data de Extração", "Data de Extracao"], None),
}

for _row in ws_ind.iter_rows(min_row=2, values_only=True):
    if not any(v not in (None, "") for v in _row):
        continue
    _rec = _selo_record_from_row(_row, _selo_headers_list, _selo_cols)
    df_selo.append(_rec)

    _ano = _rec["Ano"]
    _ies = _selo_sigla(_rec["Unidade Orçamentária"])
    _cod = _rec["Código Indicador"]
    if _ies not in IEES_PR or _ano is None:
        continue
    _k = (_ies, _ano)
    if _k not in selo_data:
        selo_data[_k] = {"notaFinal": None, "indicadores": {}}
    if _rec["Nota Final da Unidade"] is not None:
        selo_data[_k]["notaFinal"] = safe_float(_rec["Nota Final da Unidade"])
    if _rec["Nota do Bimestre"] is not None and _rec["Bimestre"]:
        _selo_bimestre_acc.setdefault((_ies, _ano, _rec["Bimestre"]), []).append(_rec["Nota do Bimestre"])
    if _cod:
        _acc = _selo_indicator_acc.setdefault(
            (_ies, _ano, _cod),
            {"weighted": 0.0, "weight": 0.0, "values": [], "eixo": None, "nome": None, "maximo": None},
        )
        _acc["eixo"] = _rec["Eixo"] or _acc["eixo"]
        _acc["nome"] = _rec["Nome Indicador"] or _acc["nome"]
        _acc["maximo"] = _rec["Nota Máxima do Indicador"] or _acc["maximo"]
        _nota_ind = _rec["Nota do Indicador no Bimestre"]
        if _nota_ind is not None:
            _peso = _rec["Peso do Bimestre"]
            if _rec["Bimestre"] and _peso is not None:
                _w = _peso if abs(_peso) <= 1 else _peso / 100
                _acc["weighted"] += _nota_ind * _w
                _acc["weight"] += _w
            else:
                _acc["values"].append(_nota_ind)

# Compatibilidade com a versão anterior do arquivo, que trazia a nota final em aba separada.
if "Nota_Final_Unidade" in wb_selo.sheetnames:
    ws_nf = wb_selo["Nota_Final_Unidade"]
    _nf_header_row = next(ws_nf.iter_rows(min_row=1, max_row=1, values_only=True), ())
    _nf_headers = _selo_headers(_nf_header_row)
    _nf_hidx = _selo_header_index(_nf_headers)
    _nf_cols = {
        "ano": _selo_col(_nf_hidx, ["Ano"], 0),
        "unidade": _selo_col(_nf_hidx, ["Unidade Orçamentária", "Unidade Orcamentaria", "IEES", "Sigla"], 2),
        "nota_final": _selo_col(_nf_hidx, ["Nota Final da Unidade", "Nota Final Unidade", "Nota Final"], 3),
    }
    for _row in ws_nf.iter_rows(min_row=2, values_only=True):
        _ano = _selo_year(_selo_cell(_row, _nf_cols["ano"]))
        _ies = _selo_sigla(_selo_cell(_row, _nf_cols["unidade"]))
        if _ies not in IEES_PR or _ano is None:
            continue
        _k = (_ies, _ano)
        if _k not in selo_data:
            selo_data[_k] = {"notaFinal": None, "indicadores": {}}
        _nota_final = safe_float(_selo_cell(_row, _nf_cols["nota_final"]))
        selo_data[_k]["notaFinal"] = _nota_final

wb_selo.close()

for (_ies, _ano, _cod), _acc in _selo_indicator_acc.items():
    if _acc["weight"] > 0:
        _nota = safe_float(_acc["weighted"], 2)
    elif _acc["values"]:
        _nota = safe_float(sum(_acc["values"]) / len(_acc["values"]), 2)
    else:
        _nota = None
    if _nota is None:
        continue
    selo_data.setdefault((_ies, _ano), {"notaFinal": None, "indicadores": {}})
    selo_data[(_ies, _ano)]["indicadores"][_cod] = _nota
    df_selo_indicadores.append({
        "Unidade Orçamentária": _ies,
        "Ano": _ano,
        "Código Indicador": _cod,
        "Nome Indicador": _acc["nome"] or _SELO_INDICADORES.get(_cod, {}).get("nome"),
        "Eixo": _acc["eixo"] or _SELO_INDICADORES.get(_cod, {}).get("eixo"),
        "Nota do Indicador": _nota,
        "Nota Máxima do Indicador": _acc["maximo"] or _SELO_INDICADORES.get(_cod, {}).get("maximo"),
    })

for (_ies, _ano, _bim), _vals in sorted(_selo_bimestre_acc.items()):
    if not _vals:
        continue
    df_selo_bimestres.append({
        "Unidade Orçamentária": _ies,
        "Ano": _ano,
        "Bimestre": _bim,
        "Nota do Bimestre": safe_float(sum(_vals) / len(_vals), 2),
    })

for (_ies, _ano), _rec in sorted(selo_data.items()):
    df_selo_resumo.append({
        "Unidade Orçamentária": _ies,
        "Ano": _ano,
        "Nota Final da Unidade": _rec.get("notaFinal"),
    })

_selo_anos = sorted({r["Ano"] for r in df_selo if r["Ano"] is not None} | {r["Ano"] for r in df_selo_resumo if r["Ano"] is not None})
_selo_unidades = sorted({_selo_sigla(r["Unidade Orçamentária"]) for r in df_selo if _selo_sigla(r["Unidade Orçamentária"])})
_selo_bimestres = sorted({r["Bimestre"] for r in df_selo if r["Bimestre"]})
_selo_codigos = sorted({r["Código Indicador"] for r in df_selo if r["Código Indicador"]})
_selo_alertas = []
_selo_warn_missing("Unidades SELO-PR", _SELO_UNIDADES_ESPERADAS, set(_selo_unidades), _selo_alertas)
_selo_warn_missing("Anos SELO-PR", _SELO_ANOS_ESPERADOS, set(_selo_anos), _selo_alertas)
if _SELO_BIMESTRES_ESPERADOS:
    _selo_warn_missing("Bimestres SELO-PR", _SELO_BIMESTRES_ESPERADOS, set(_selo_bimestres), _selo_alertas)
_selo_warn_missing("Indicadores SELO-PR", _SELO_INDICADORES_ESPERADOS, set(_selo_codigos), _selo_alertas)
_selo_linhas_esperadas = (
    len(_SELO_UNIDADES_ESPERADAS)
    * len(_SELO_ANOS_ESPERADOS)
    * len(_SELO_INDICADORES_ESPERADOS)
)
if len(df_selo) != _selo_linhas_esperadas:
    _selo_alertas.append(
        f"Total de linhas fora do esperado para base anual ({_selo_linhas_esperadas}): {len(df_selo)}"
    )

selo_diagnostico = {
    "arquivo": _SELO_FILE,
    "caminho": str(caminho_selo),
    "aba": _selo_sheet,
    "linhas": len(df_selo),
    "colunas": len(_selo_headers_list),
    "anos": _selo_anos,
    "unidades": _selo_unidades,
    "bimestres": _selo_bimestres,
    "indicadores": _selo_codigos,
    "alertas": _selo_alertas,
    "preview": [{k: v for k, v in row.items() if k != "_raw"} for row in df_selo[:20]],
}

for _ies in IEES_PR:
    _anos_ies = sorted(ano for (sigla, ano) in selo_data if sigla == _ies)
    if not _anos_ies:
        continue
    _ano_ref = _anos_ies[-1]
    _nota_ref = selo_data.get((_ies, _ano_ref), {}).get("notaFinal")
    _key = _ies.lower()
    if _key in results:
        results[_key]["seloNotaFinal"] = _nota_ref
        sources[_key]["seloNotaFinal"] = (
            f"{_SELO_FILE} / {_selo_sheet}"
            f" / Nota Final da Unidade / ano={_ano_ref}"
        )

# 16c. Validação no stderr
_SEP16 = "─" * 88
print("", file=sys.stderr)
print("====================================", file=sys.stderr)
print("✅ Base SELO-PR carregada com sucesso", file=sys.stderr)
print(f"Arquivo SELO-PR: {caminho_selo}", file=sys.stderr)
print(f"Linhas SELO-PR: {len(df_selo)}", file=sys.stderr)
print(f"Colunas SELO-PR: {len(_selo_headers_list)}", file=sys.stderr)
print(f"Anos SELO-PR: {_selo_anos}", file=sys.stderr)
print(f"Unidades SELO-PR: {_selo_unidades}", file=sys.stderr)
print(f"Bimestres SELO-PR: {_selo_bimestres}", file=sys.stderr)
print(f"Indicadores SELO-PR: {_selo_codigos}", file=sys.stderr)
for _alerta in _selo_alertas:
    print(f"⚠️ ALERTA SELO-PR: {_alerta}", file=sys.stderr)
print("====================================", file=sys.stderr)
print(_SEP16, file=sys.stderr)
print("Seção 16 — SELO-PR | Notas anuais por IES (0–100)", file=sys.stderr)
print(f"Fonte: {_SELO_FILE} / {_selo_sheet}", file=sys.stderr)
print(_SEP16, file=sys.stderr)
print(f"{'IES':<12} {'Ano':>5} {'Final':>6}  {'1.1':>4} {'1.2':>4} {'1.3':>4} {'1.4':>4} | {'2.1':>4} {'2.2':>4} {'2.3':>4} | {'3.1':>4} {'3.2':>4} {'3.3':>4} {'3.4':>4}", file=sys.stderr)
print(_SEP16, file=sys.stderr)
for _ies in IEES_PR:
    for _ano in _selo_anos:
        _kk = (_ies, _ano)
        _r = selo_data.get(_kk)
        if not _r:
            print(f"{_ies:<12} {_ano:>5} {'N/D':>6}", file=sys.stderr)
        else:
            _inds = _r["indicadores"]
            def _gi(cod): return _inds.get(cod) or 0
            print(
                f"{_ies:<12} {_ano:>5} {(_r['notaFinal'] or 0):>6.1f}"
                f"  {_gi('1.1'):>4.1f} {_gi('1.2'):>4.1f} {_gi('1.3'):>4.1f} {_gi('1.4'):>4.1f}"
                f" | {_gi('2.1'):>4.1f} {_gi('2.2'):>4.1f} {_gi('2.3'):>4.1f}"
                f" | {_gi('3.1'):>4.1f} {_gi('3.2'):>4.1f} {_gi('3.3'):>4.1f} {_gi('3.4'):>4.1f}",
                file=sys.stderr,
            )
print(_SEP16, file=sys.stderr)


# ── Seção 17 — Cluster Específico (C1–C8) ─────────────────────────────────────
# Fonte: Clusterização específica.xlsx / Planilha1
# Colunas: [0]=Cluster (C1..C8) [1]=Nome do cluster [2]=Perfil do grupo
#          [3]=Variáveis consideradas [4]=IEES (siglas separadas por vírgula)
# Atribuição estática (curadoria manual do Instituto Publix), não recalculada por ano.
# Cobre as 40 IES (7 PR + 33 BR); disponível nos escopos Paraná e Brasil.
# Segue o mesmo padrão de clusters_raw usado para V1–V8: a chave "c_especifico"
# entra em clusters_raw[sigla] e flui naturalmente para precomputed["clusters"].

wb_ce = openpyxl.load_workbook(
    DATA_DIR / "Clusterização específica.xlsx", read_only=True, data_only=True
)
ws_ce = wb_ce["Planilha1"]
next(ws_ce.iter_rows(min_row=1, max_row=1))  # skip header

clusters_especificos_catalog = []
_sigla_to_cluster = {}

for row in ws_ce.iter_rows(min_row=2, values_only=True):
    codigo = row[0] if row else None
    if not codigo or not str(codigo).strip().upper().startswith("C"):
        continue  # ignora linha de nota metodológica no rodapé da planilha
    codigo = str(codigo).strip()
    nome    = str(row[1]).strip() if len(row) > 1 and row[1] else None
    perfil  = str(row[2]).strip() if len(row) > 2 and row[2] else None
    varis   = str(row[3]).strip() if len(row) > 3 and row[3] else None
    ies_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""
    siglas  = [s.strip() for s in ies_str.split(",") if s.strip()]

    clusters_especificos_catalog.append({
        "codigo": codigo,
        "nome": nome,
        "perfil": perfil,
        "variaveis": varis,
        "ies": siglas,
    })
    for s in siglas:
        _sigla_to_cluster[s.upper()] = {"codigo": codigo, "nome": nome, "perfil": perfil}

wb_ce.close()

_SRC_CE = "Clusterização específica.xlsx / Planilha1 / atribuição estática C1–C8"
for iees in IEES:
    key = iees.lower()
    info = _sigla_to_cluster.get(iees.upper())
    if info is None:
        continue
    label = f"{info['codigo']} - {info['nome']}"
    results[key]["clusterEspecificoCodigo"] = info["codigo"]
    results[key]["clusterEspecificoPerfil"] = info["perfil"]
    sources[key]["clusterEspecificoCodigo"] = _SRC_CE
    if iees in clusters_raw:
        clusters_raw[iees]["c_especifico"] = label
    else:
        clusters_raw[iees] = {"c_especifico": label}

# Validação — confere se as 40 IES foram classificadas
_nao_classificadas = [i for i in IEES if i.upper() not in _sigla_to_cluster]
if _nao_classificadas:
    print(
        f"[AVISO] IES sem Cluster Específico atribuído: {', '.join(_nao_classificadas)}",
        file=sys.stderr,
    )


# ── 13. Referência Geral — melhor valor bruto entre as IES (whitelist v1) ────
# Substitui, na v1 do frontend, a "média do cluster" (que reage a filtro) por
# um valor FIXO: o maior (ou menor, conforme polaridade) valor bruto de uma
# única IES entre o universo aplicável — nunca recalculado a partir de linhas
# filtradas. Mesmo princípio já usado em benchmark_cine (Seção 2c).
#
# Roda por último (depois de todas as seções) porque a whitelist inclui campos
# de seções tardias: doctors (1), occupancy/dropout/completion (2),
# facultyOcc/docTaxaUtil/cres/docCresOciosidade/docCresPartic/docTidePartic (3),
# cnpq (4), pctExcelencia (5b), employment/salary (8), insertionRatePR (9),
# tx_execucao_empenho/grau_contingenciamento/tx_liquidacao/tx_pagamento_liq
# (6b — merge de d8050_by_year 2024 em results, ~linha 1594).
#
# Excluídos da v1 (decisão desta rodada — não implementados aqui):
#   - Indicadores "Neutra" no catálogo: var_dotacao_loa, part_pessoal,
#     part_outras_correntes, razão correntes/capital, razaoDocenteDiscente.
#   - Índices sintéticos sem entrada própria: academicPerformanceIndex,
#     costPerStudent, costEquivalentStudent, dimensionScore, cnpqLinks,
#     cboDistribution, occupationalDiversity, foreignFacultyRate,
#     mobilityRate, pgForeignShare, pgProductivityShare.
#   - Campos sem polaridade/universo confirmado: capes (bruto), vinculos,
#     egressosMunicipios, docVagasTotais/docVagasDisp/docVagasOcupadas/
#     docVagasCond, docChMedia.
#   - Composição orçamentária sem entrada isolada: part_capital,
#     part_recursos_livres, part_fonte_500/501, part_demais_vincul,
#     transfers, freeResources, ownResources.
#   - DIVERGÊNCIA ENCONTRADA NESTA RODADA — cbo2Rate (IND-39): a whitelist
#     pedia a inclusão deste campo, mas `grep -r ind39 pipeline/` não retorna
#     NENHUMA ocorrência em todo o pipeline (nem em assemble_final.py, nem em
#     nenhum enrich_*.py), e o JSON gerado não tem a chave "ind39" em nenhum
#     byYear de nenhuma IES (confirmado lendo data/seti_precomputed.json).
#     Ou seja: o campo `employmentMetrics(u).cbo2Rate` do frontend (que lê
#     `real.ind39 ?? fórmula sintética`) sempre cai no fallback sintético hoje
#     — não há dado real precomputado para calcular uma referência geral.
#     NÃO implementado; card de cbo2Rate (Aba 7) continua com média do
#     cluster, igual aos demais índices sintéticos excluídos.
_REFERENCIA_GERAL_WHITELIST = {
    # campo:                    (universo, polaridade, rótulo do universo)
    "dropout":                 (IEES,    "menor", "40 IES"),
    "completion":              (IEES,    "maior", "40 IES"),
    "doctors":                 (IEES,    "maior", "40 IES"),
    "occupancy":               (IEES,    "maior", "40 IES"),
    "cnpq":                    (IEES,    "maior", "40 IES"),
    "pctExcelencia":           (IEES,    "maior", "40 IES"),
    "facultyOcc":              (IEES_PR, "maior", "7 IES-PR"),
    "docTaxaUtil":             (IEES_PR, "maior", "7 IES-PR"),
    "cres":                    (IEES_PR, "maior", "7 IES-PR"),
    "docCresOciosidade":       (IEES_PR, "menor", "7 IES-PR"),
    "docCresPartic":           (IEES_PR, "maior", "7 IES-PR"),
    "docTidePartic":           (IEES_PR, "maior", "7 IES-PR"),
    "employment":              (IEES_PR, "maior", "7 IES-PR"),
    "insertionRatePR":         (IEES_PR, "maior", "7 IES-PR"),
    "salary":                  (IEES_PR, "maior", "7 IES-PR"),
    "tx_execucao_empenho":     (IEES_PR, "maior", "7 IES-PR"),
    "grau_contingenciamento":  (IEES_PR, "menor", "7 IES-PR"),
    "tx_liquidacao":           (IEES_PR, "maior", "7 IES-PR"),
    "tx_pagamento_liq":        (IEES_PR, "maior", "7 IES-PR"),
}


def calcular_referencia_geral():
    resultado = {}
    for campo, (universo, polaridade, universo_label) in _REFERENCIA_GERAL_WHITELIST.items():
        melhor_valor = None
        melhor_sigla = None
        for iees in universo:
            v = results.get(iees.lower(), {}).get(campo)
            if v is None:
                continue
            if (
                melhor_valor is None
                or (polaridade == "maior" and v > melhor_valor)
                or (polaridade == "menor" and v < melhor_valor)
            ):
                melhor_valor, melhor_sigla = v, iees
        resultado[campo] = {
            "valor": melhor_valor,
            "sigla": melhor_sigla,
            "universo": universo_label,
            "polaridade": polaridade,
        }
    return resultado


referencia_geral = calcular_referencia_geral()


# ── Saída stdout (retrocompatível) ────────────────────────────────────────────

print(json.dumps({"results": results, "sources": sources}, indent=2, ensure_ascii=False))

# ── Salva seti_precomputed.json para o dashboard ──────────────────────────────
# Formato: {year, indicators, sources, clusters, quartiRefs, byYear, generated}
# clusters:   {SIGLA: {v1..v8: label_str}} — lido da Estratificação, nunca estático
# quartiRefs: lista com limiares e rótulos de cada variável de agrupamento
# byYear:     {SIGLA: {str(ano): {despesa8050_fields}}} para 2024/2025/2026

precomputed = {
    "generated": datetime.datetime.now().isoformat(timespec="seconds"),
    "year": "2024",
    "indicators": {iees: results[iees.lower()] for iees in IEES},
    "sources":    {iees: sources[iees.lower()] for iees in IEES},
    "clusters":   {iees: clusters_raw.get(iees, {}) for iees in IEES},
    "quartiRefs": quartis_ref,
    "clustersEspecificos": clusters_especificos_catalog,
    "composicaoFontes": composicaoFontes,
    "cursosDetalhado": cursos_detalhado,
    # cursosDetalhadoByYear: {SIGLA: {str(ano): [grupos]}} para 2020-2024 —
    # Round 3a, pré-requisito para o front-end (Round 3b) combinar Tipo de
    # Curso/Modalidade/Grande Área com o filtro de Ano. cursosDetalhado
    # (acima) continua sendo exportado sem alteração nesta rodada.
    "cursosDetalhadoByYear": cursos_detalhado_by_year,
    # benchmarkCine: {area: {dropout: {pr, br, referencia, origem}, occupancyTipo: {...}}}
    # pendência de validação metodológica (ponderação e escopo do campo occupancy) — ver Seção 2c
    "benchmarkCine": benchmark_cine,
    # referenciaGeral: {campo: {valor, sigla, universo, polaridade}} — melhor
    # valor bruto de uma única IES (fixo, não reage a filtro). Whitelist v1 —
    # ver Seção 13. cbo2Rate (IND-39) não incluído: sem dado real precomputado.
    "referenciaGeral": referencia_geral,
    # byYear: 2024 = todos os indicadores para as 40 IES;
    # 2025/2026 = apenas campos D8050 para as 7 IES-PR
    "byYear": {
        iees: {
            "2024": results[iees.lower()],
            **{
                yr: d8050_by_year.get(iees, {}).get(yr, {})
                for yr in ["2025", "2026"]
                if d8050_by_year.get(iees, {}).get(yr)
            },
        }
        for iees in IEES
    },
}
# seloData: {SIGLA: {str(ano): {"notaFinal": float, "indicadores": {"1.1": float, ...}}}}
_selo_export = {}
for _ies in IEES_PR:
    _selo_export[_ies] = {}
    for _ano in _selo_anos:
        _kk = (_ies, _ano)
        _r = selo_data.get(_kk)
        if _r:
            _selo_export[_ies][str(_ano)] = _r

precomputed["seloData"]           = _selo_export
precomputed["seloIndicadores"]    = _SELO_INDICADORES
precomputed["seloDiagnostico"]    = selo_diagnostico
precomputed["seloResumo"]         = df_selo_resumo
precomputed["seloBimestres"]      = df_selo_bimestres
precomputed["seloIndicadorNotas"] = df_selo_indicadores

json_path = DATA_DIR / "seti_precomputed.json"
with open(json_path, "w", encoding="utf-8") as _f:
    json.dump(precomputed, _f, indent=2, ensure_ascii=False)
print(f"[OK] {json_path}", file=sys.stderr)

# ── Validação Seção 9 — comparação com valores de referência 2024 ─────────────

_REF_2024 = {
    "UEL":      {"dotacao_inicial": 641.6, "orcamento_atualizado": 753.7, "liquidado": 694.6, "part_pessoal": 82.6, "tx_execucao_empenho": 94.7},
    "UEM":      {"dotacao_inicial": 708.8, "orcamento_atualizado": 871.6, "liquidado": 762.1, "part_pessoal": 77.6, "tx_execucao_empenho": 91.4},
    "UEPG":     {"dotacao_inicial": 348.2, "orcamento_atualizado": 423.7, "liquidado": 399.4, "part_pessoal": 83.0, "tx_execucao_empenho": 94.8},
    "UNIOESTE": {"dotacao_inicial": 468.0, "orcamento_atualizado": 561.3, "liquidado": 520.4, "part_pessoal": 81.5, "tx_execucao_empenho": 94.8},
    "UNICENTRO":{"dotacao_inicial": 311.7, "orcamento_atualizado": 380.6, "liquidado": 341.0, "part_pessoal": 79.6, "tx_execucao_empenho": 92.5},
    "UENP":     {"dotacao_inicial": 130.8, "orcamento_atualizado": 164.9, "liquidado": 155.9, "part_pessoal": 80.8, "tx_execucao_empenho": 95.8},
    "UNESPAR":  {"dotacao_inicial": 258.6, "orcamento_atualizado": 319.8, "liquidado": 302.8, "part_pessoal": 85.0, "tx_execucao_empenho": 97.0},
}

_VAL_COLS = [
    ("dotacao_inicial",      "Dot.Ini(R$M)"),
    ("orcamento_atualizado", "Orç.Atu(R$M)"),
    ("liquidado",            "Liquid.(R$M)"),
    ("part_pessoal",         "%Pessoal"),
    ("tx_execucao_empenho",  "%Execução"),
]

_W_IES = 10
_W_COL = 15
_SEP = "─" * (_W_IES + _W_COL * len(_VAL_COLS))

print("", file=sys.stderr)
print(_SEP, file=sys.stderr)
print("Seção 9 — Despesa 8050 | Validação ano 2024 (extraído vs referência)", file=sys.stderr)
print(_SEP, file=sys.stderr)
hdr_line = f"{'IES':<{_W_IES}}" + "".join(f"{lbl:>{_W_COL}}" for _, lbl in _VAL_COLS)
print(hdr_line, file=sys.stderr)
print(_SEP, file=sys.stderr)

all_ok = True
for sigla in IEES_PR:
    yr_data = d8050_by_year.get(sigla, {}).get("2024", {})
    ref     = _REF_2024.get(sigla, {})
    row_out = f"{sigla:<{_W_IES}}"
    for field, _ in _VAL_COLS:
        got = yr_data.get(field)
        exp = ref.get(field)
        if got is None:
            cell = "N/D"
        elif exp is None:
            cell = f"{got:.1f}"
        else:
            diff = abs(got - exp)
            ok_mark = "OK" if diff <= 1.0 else "!!"
            if ok_mark == "!!":
                all_ok = False
            cell = f"{got:.1f}({ok_mark})"
        row_out += f"{cell:>{_W_COL}}"
    print(row_out, file=sys.stderr)

print(_SEP, file=sys.stderr)
ref_hdr = f"{'(ref)':<{_W_IES}}" + "".join(f"{lbl:>{_W_COL}}" for _, lbl in _VAL_COLS)
print(ref_hdr, file=sys.stderr)
for sigla in IEES_PR:
    ref = _REF_2024.get(sigla, {})
    row_out = f"{sigla:<{_W_IES}}"
    for field, _ in _VAL_COLS:
        exp = ref.get(field)
        row_out += f"{exp:>{_W_COL}.1f}" if exp is not None else f"{'?':>{_W_COL}}"
    print(row_out, file=sys.stderr)

print(_SEP, file=sys.stderr)
status = "PASS — todos dentro da tolerância (±1)" if all_ok else "ATENÇÃO — diferenças acima de ±1 detectadas"
print(f"Status: {status}", file=sys.stderr)
print(_SEP, file=sys.stderr)

# ── Tabela de resumo (stderr) ─────────────────────────────────────────────────

header = f"{'Indicador':<20}" + "".join(f"{i:<13}" for i in IEES)
print("", file=sys.stderr)
print("=" * (20 + 13 * len(IEES)), file=sys.stderr)
print(header, file=sys.stderr)
print("=" * (20 + 13 * len(IEES)), file=sys.stderr)
for ind in INDICATORS:
    row_str = f"{ind:<20}"
    for iees in IEES:
        v = results[iees.lower()].get(ind)
        row_str += f"{str(v):<13}"
    print(row_str, file=sys.stderr)
print("=" * (20 + 13 * len(IEES)), file=sys.stderr)

# ── Tabela 5b — validação CAPES Discentes / Docentes (7 IES-PR) ──────────────
_5B_COLS = [
    ("pgTop",               "pgTop"),
    ("capes",               "capes"),
    ("pgMestrado",          "pgMest"),
    ("pgMestradoProf",      "pgProf"),
    ("pgDoutorado",         "pgDout"),
    ("pgMunicipiosDistintos", "pgMuni"),
    ("pctExcelencia",       "pctExc%"),
    ("discMestrado",        "discMest"),
    ("discDoutorado",       "discDout"),
    ("tituladosMestrado",   "titMest"),
    ("tituladosDoutorado",  "titDout"),
    ("docPermanentes",      "docPerm"),
    ("docColaboradores",    "docColab"),
    ("docVisitantes",       "docVis"),
    ("razaoDocenteDiscente", "disc/docP"),
]
_W5B_IES = 12
_W5B_COL = 10
_SEP5B = "─" * (_W5B_IES + _W5B_COL * len(_5B_COLS))
print("", file=sys.stderr)
print(_SEP5B, file=sys.stderr)
print("Seção 5b — CAPES Discentes/Docentes | Validação 7 IES-PR", file=sys.stderr)
print("  pgTop/capes: CD_CONCEITO_CURSO (Base_Cursos)  |  pctExc%: CD_CONCEITO_PROGRAMA>=6 (Base_Discentes)", file=sys.stderr)
print(_SEP5B, file=sys.stderr)
_hdr5b = f"{'IES':<{_W5B_IES}}" + "".join(f"{_lbl:>{_W5B_COL}}" for _, _lbl in _5B_COLS)
print(_hdr5b, file=sys.stderr)
print(_SEP5B, file=sys.stderr)
for _sig5b in IEES_PR:
    _k5b   = _sig5b.lower()
    _row5b = f"{_sig5b:<{_W5B_IES}}"
    for _fld5b, _ in _5B_COLS:
        _v5b = results[_k5b].get(_fld5b)
        if _v5b is None:
            _cell5b = "N/D"
        elif isinstance(_v5b, float):
            _cell5b = f"{_v5b:.1f}"
        elif isinstance(_v5b, dict):
            _cell5b = f"n={len(_v5b)}"
        else:
            _cell5b = str(_v5b)
        _row5b += f"{_cell5b:>{_W5B_COL}}"
    print(_row5b, file=sys.stderr)
print(_SEP5B, file=sys.stderr)

# ── Relatório de rastreabilidade (pipeline_report.json) ──────────────────────
_pipeline_end = datetime.datetime.now()
_duration = round((_pipeline_end - _pipeline_start).total_seconds(), 1)

_ind_summary = {}
for ind in INDICATORS:
    _ind_summary[ind] = {
        "values": {iees: results[iees.lower()].get(ind) for iees in IEES},
        "sources": {iees: sources[iees.lower()].get(ind) for iees in IEES_PR},
        "null_count_pr":  sum(1 for iees in IEES_PR if results[iees.lower()].get(ind) is None),
        "null_count_br":  sum(1 for iees in IEES_BR if results[iees.lower()].get(ind) is None),
    }

_report = {
    "run_at":            _pipeline_start.isoformat(timespec="seconds"),
    "finished_at":       _pipeline_end.isoformat(timespec="seconds"),
    "duration_seconds":  _duration,
    "output_json":       str(DATA_DIR / "seti_precomputed.json"),
    "ies_pr":            IEES_PR,
    "ies_br":            IEES_BR,
    "total_indicators":  len(INDICATORS),
    "alerts":            _pipeline_alerts + [f"SELO-PR: {a}" for a in _selo_alertas],
    "indicators":        _ind_summary,
}

_report_path = DATA_DIR / "pipeline_report.json"
with open(_report_path, "w", encoding="utf-8") as _rf:
    json.dump(_report, _rf, ensure_ascii=False, indent=2, default=str)
print(f"\n[OK] Relatório de rastreabilidade: {_report_path}", file=sys.stderr)
print(f"     Duração total: {_duration}s | {len(INDICATORS)} indicadores | alertas: {len(_report['alerts'])}", file=sys.stderr)
