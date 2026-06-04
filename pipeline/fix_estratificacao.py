"""
Corrige Estratificação_IES_Estaduais_BR.xlsx:
  - Remove UNIVESP de todas as sheets
  - Corrige URCA: students 57628 → 9659
  - Recalcula quartis V1 e V2 com 40 IES
  - Atualiza faixas de porte (V1) para 6 IES que mudam de grupo
  - Renumera e atualiza títulos/notas
"""
import sys, shutil, statistics, copy
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

DATA_DIR = Path(__file__).parent.parent / "data"
XLSX_PATH = DATA_DIR / "Estratificação_IES_Estaduais_BR.xlsx"
BACKUP_PATH = DATA_DIR / "Estratificação_IES_Estaduais_BR.BACKUP.xlsx"

# ── Backup ───────────────────────────────────────────────────────────────────
shutil.copy2(XLSX_PATH, BACKUP_PATH)
print(f"Backup: {BACKUP_PATH}")

# ── Novos quartis V1 (calculados sem UNIVESP, URCA=9659) ─────────────────────
# Q1=25242, Med=50410, Q3=85987
NEW_V1 = {"q1": 25242, "med": 50410, "q3": 85987,
           "label_q1": "Pequeno Porte", "label_q2": "Médio-Pequeno Porte",
           "label_q3": "Médio-Grande Porte", "label_q4": "Grande Porte"}

# Novos quartis V2 (sem UNIVESP): Q1=162, Med=320, Q3=573
NEW_V2 = {"q1": 162, "med": 320, "q3": 573,
           "label_q1": "Oferta Reduzida", "label_q2": "Oferta Moderada",
           "label_q3": "Oferta Ampla", "label_q4": "Oferta Extensa"}

def v1_faixa(students):
    if students is None: return None
    if students <= NEW_V1["q1"]:  return NEW_V1["label_q1"]
    if students <= NEW_V1["med"]: return NEW_V1["label_q2"]
    if students <= NEW_V1["q3"]:  return NEW_V1["label_q3"]
    return NEW_V1["label_q4"]

def v2_faixa(courses):
    if courses is None: return None
    if courses <= NEW_V2["q1"]:  return NEW_V2["label_q1"]
    if courses <= NEW_V2["med"]: return NEW_V2["label_q2"]
    if courses <= NEW_V2["q3"]:  return NEW_V2["label_q3"]
    return NEW_V2["label_q4"]

# ── Carrega workbook em modo escrita ─────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 0 — Referência de Quartis
# ═══════════════════════════════════════════════════════════════════════════
ws0 = wb["0_Referência de Quartis"]

for row in ws0.iter_rows():
    for cell in row:
        v = cell.value
        if not isinstance(v, str): continue
        # V1 row: "Porte Institucional"
        if "Porte Institucional" in v and "Matr" in v:
            r = cell.row
            ws0.cell(r, 3).value = "25.242"   # Q1
            ws0.cell(r, 4).value = "50.410"   # Mediana
            ws0.cell(r, 5).value = "85.987"   # Q3
            print(f"[Sheet 0] V1 quartis atualizados: Q1=25.242, Med=50.410, Q3=85.987")
        # V2 row: "Oferta de Cursos"
        if "Oferta de Cursos" in v:
            r = cell.row
            ws0.cell(r, 3).value = "162"
            ws0.cell(r, 4).value = "320"
            ws0.cell(r, 5).value = "573"
            print(f"[Sheet 0] V2 quartis atualizados: Q1=162, Med=320, Q3=573")
        # Nota metodológica: 41 → 40
        if "41 IES" in v:
            cell.value = v.replace("41 IES", "40 IES")
            print(f"[Sheet 0] Nota atualizada: '41 IES' → '40 IES'")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — Matriz de Estratificação
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb["1_Matriz de Estratificação"]

# Atualiza título
for row in ws1.iter_rows(min_row=1, max_row=2):
    for cell in row:
        if isinstance(cell.value, str) and "41 IES" in cell.value:
            cell.value = cell.value.replace("41 IES", "40 IES")
            print(f"[Sheet 1] Título atualizado: '41 IES' → '40 IES'")

# Encontra linha da UNIVESP
univesp_row = None
urca_row = None
for row in ws1.iter_rows(min_row=6):
    c = row[2].value  # coluna C = Sigla
    if c == "UNIVESP":
        univesp_row = row[0].row
    if c == "URCA":
        urca_row = row[0].row

print(f"[Sheet 1] UNIVESP na linha {univesp_row}, URCA na linha {urca_row}")

# Corrige URCA (antes de deletar UNIVESP para não deslocar)
if urca_row:
    ws1.cell(urca_row, 7).value = 9659   # G: Nº Matrículas
    ws1.cell(urca_row, 8).value = 9659   # H: Total Matrículas (Valor)
    ws1.cell(urca_row, 9).value = v1_faixa(9659)  # I: Faixa de Porte
    print(f"[Sheet 1] URCA: students=9659, faixa='{v1_faixa(9659)}'")

# Atualiza V1 faixa das outras 5 IES que mudam
faixa_changes = {
    "UECE":      v1_faixa(86904),   # Médio-Grande → Grande Porte
    "UEPA":      v1_faixa(52551),   # Médio-Pequeno → Médio-Grande
    "UNIMONTES": v1_faixa(51465),   # Médio-Pequeno → Médio-Grande
    "UEMS":      v1_faixa(33243),   # Pequeno → Médio-Pequeno
    "UNEAL":     v1_faixa(28924),   # Pequeno → Médio-Pequeno
}
for row in ws1.iter_rows(min_row=6):
    sig = row[2].value
    if sig in faixa_changes:
        old = row[8].value
        ws1.cell(row[0].row, 9).value = faixa_changes[sig]
        print(f"[Sheet 1] {sig}: faixa '{old}' → '{faixa_changes[sig]}'")

# Deleta linha da UNIVESP
ws1.delete_rows(univesp_row, 1)
print(f"[Sheet 1] UNIVESP deletada (linha {univesp_row})")

# Renumera coluna A (#): rows 6..45 → 1..40
for i, row in enumerate(ws1.iter_rows(min_row=6), start=1):
    sig = row[2].value
    if sig and str(sig).strip() and str(sig).strip() != "Sigla":
        row[0].value = i
if i != 40:
    print(f"  AVISO: Renumeração terminou em {i}, esperado 40")
else:
    print(f"[Sheet 1] Renumerado 1-40 OK")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — Porte Institucional (reorganização V1)
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb["2_Porte Institucional"]

# Lê todos os dados das 4 seções (sigla→row data)
s2_rows = {}
for row in ws2.iter_rows(min_row=6):
    sig = row[2].value
    if not sig or str(sig).strip() in ('', 'Sigla', '#'): continue
    sig = str(sig).strip()
    s2_rows[sig] = {
        "uf":        row[1].value,
        "nome":      row[3].value,
        "mun":       row[4].value,
        "students":  row[5].value,
        "vagas":     row[7].value,
        "row_idx":   row[0].row,
    }

# Corrige URCA nos dados em memória
s2_rows["URCA"]["students"] = 9659

# Define novos grupos (ordenados por students desc)
def make_group(siglas):
    items = [(s, s2_rows[s]) for s in siglas if s in s2_rows]
    return sorted(items, key=lambda x: (x[1]["students"] or 0), reverse=True)

new_q1 = make_group(["UENF","UENP","UERGS","UNITINS","UEMASUL","UEAP","URCA","UNCISAL","UERR","UnDF"])
new_q2 = make_group(["UERN","UEFS","UEPG","UNESPAR","UESB","UVA","UESC","UNICENTRO","UEMS","UNEAL"])
new_q3 = make_group(["UEA","UEPB","UPE","UEG","UEM","UEL","UNIOESTE","UDESC","UEPA","UNIMONTES"])
new_q4 = make_group(["USP","UNESP","UERJ","UEMG","UEMA","UNICAMP","UNEMAT","UNEB","UESPI","UECE"])

# Escreve grupos nas suas posições fixas de linhas
def write_group(ws, data_rows_start, faixa_label, items):
    """Sobrescreve as linhas de dados do grupo (sem alterar headers)."""
    for i, (sig, d) in enumerate(items):
        r = data_rows_start + i
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = d["uf"]
        ws.cell(r, 3).value = sig
        ws.cell(r, 4).value = d["nome"]
        ws.cell(r, 5).value = d["mun"]
        ws.cell(r, 6).value = d["students"]
        ws.cell(r, 7).value = faixa_label
        ws.cell(r, 8).value = d["vagas"]

# Current structure in sheet 2:
# Q1 header=row5, colheader=row6, data=rows7-17(11 rows), blank=18
# Q2 header=row19, colheader=row20, data=rows21-30(10 rows), blank=31
# Q3 header=row32, colheader=row33, data=rows34-43(10 rows), blank=44
# Q4 header=row45, colheader=row46, data=rows47-56(10 rows)

# Q1: write 10 IES (was 11), blank out row 17
write_group(ws2, 7, NEW_V1["label_q1"], new_q1)
# Row 17 used to be the 11th Q1 IES → blank it (now it's a separator)
for col in range(1, 9):
    ws2.cell(17, col).value = None

# Q2: rows 21-30
write_group(ws2, 21, NEW_V1["label_q2"], new_q2)

# Q3: rows 34-43
write_group(ws2, 34, NEW_V1["label_q3"], new_q3)

# Q4: rows 47-56
write_group(ws2, 47, NEW_V1["label_q4"], new_q4)

# Update section headers
for row in ws2.iter_rows():
    for cell in row:
        v = cell.value
        if not isinstance(v, str): continue
        if "Pequeno Porte" in v and "IES" in v and "Q1" in v:
            cell.value = v.replace("11 IES", "10 IES")
        if "41 IES" in v: cell.value = v.replace("41 IES", "40 IES")

print(f"[Sheet 2] Grupos V1 reorganizados: Q1=10, Q2=10, Q3=10, Q4=10")

# ═══════════════════════════════════════════════════════════════════════════
# SHEETS 3-6 — Remove UNIVESP, atualiza contagem do grupo
# ═══════════════════════════════════════════════════════════════════════════
sheet_info = {
    "3_Oferta de Cursos":        ("Q1", "Oferta Reduzida",       "11 IES", "10 IES"),
    "4_Oferta Territorial":      ("Q4", "Muito Alta Dispersão",  "10 IES",  "9 IES"),
    "5_Qualificação Docente":    ("Q4", "Qualif. Consolidada",   "10 IES",  "9 IES"),
    "6_Estrutura Acadêmica":     ("Q4", "Estrutura Madura",      "10 IES",  "9 IES"),
}

for sname, (qgroup, qlabel, old_count, new_count) in sheet_info.items():
    ws = wb[sname]
    # Find and delete UNIVESP
    univesp_r = None
    for row in ws.iter_rows():
        if any(c.value == "UNIVESP" for c in row):
            univesp_r = row[0].row
            break
    if univesp_r:
        ws.delete_rows(univesp_r, 1)
        print(f"[{sname}] UNIVESP deletada (linha {univesp_r})")
    # Update section header count for the affected group
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and qlabel in v and old_count in v:
                cell.value = v.replace(old_count, new_count)
                print(f"[{sname}] Header '{qlabel}': '{old_count}' → '{new_count}'")

# ── Salva ────────────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"\n[OK] Salvo: {XLSX_PATH}")
print(f"     Backup: {BACKUP_PATH}")
