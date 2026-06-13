#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrai captação E número de vínculos de fomento do CNPq, por IES e ano, para
TODAS as 40 IES (a tabela embutida CNPQ_DATA no painel cobre só as 7 do PR;
as nacionais ficavam com vínculos nulos). Grava em byYear como cnpqCaptacao
e cnpqVinculos — nomes que o byYear() do painel já consome.

Reaproveita o dicionário CNPQ_MATCH do assemble_final.py (fonte única do
matching instituição→sigla), evitando divergência de regras.

Uso:  python3 pipeline/enrich_cnpq.py
"""
import json
import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
JSON_PATH = DATA_DIR / "seti_precomputed.json"

# ── Reusa CNPQ_MATCH do assemble_final (single source of truth) ──────────────
_src = (ROOT / "pipeline" / "assemble_final.py").read_text(encoding="utf-8")
_m = re.search(r"CNPQ_MATCH = \{.*?\n\}", _src, re.S)
_ns = {}
exec(_m.group(0), _ns)
CNPQ_MATCH = _ns["CNPQ_MATCH"]
print(f"CNPQ_MATCH: {len(CNPQ_MATCH)} IES")


def _norm(s):
    return unicodedata.normalize("NFKD", str(s).upper()).encode("ascii", "ignore").decode("ascii")


wb = openpyxl.load_workbook(DATA_DIR / "Base CNPq - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["Base_CNPq_BR"]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
col = {h: i for i, h in enumerate(headers) if h is not None}
c_inst = col["01_Instituição"]
c_ano = col["Ano"]
c_capt = col["Captação de recursos para pesquisa"]
c_vinc = col["Número de vínculos de fomento do CNPq"]

acc = {}  # (sigla, ano) -> [captacao, vinculos]
for row in ws.iter_rows(min_row=2, values_only=True):
    inst = row[c_inst]
    if not inst or not isinstance(inst, str):
        continue
    up = _norm(inst)
    for sigla, match in CNPQ_MATCH.items():
        if match(up):
            try:
                ano = int(row[c_ano])
            except (TypeError, ValueError):
                continue
            capt = float(row[c_capt]) if row[c_capt] else 0.0
            vinc = float(row[c_vinc]) if row[c_vinc] else 0.0
            a = acc.setdefault((sigla, ano), [0.0, 0.0])
            a[0] += capt
            a[1] += vinc
            break
wb.close()
print(f"CNPq: {len(acc)} pares IES×ano")

data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
by_year = data.setdefault("byYear", {})
sources = data.setdefault("sources", {})
SRC = "Base CNPq - Brasil.xlsx / Base_CNPq_BR / Σ por instituição e ano"

added = 0
for (sigla, ano), (capt, vinc) in acc.items():
    slot = by_year.setdefault(sigla, {}).setdefault(str(ano), {})
    capt_m = round(capt / 1e6, 3)
    vinc_i = int(round(vinc))
    if slot.get("cnpqCaptacao") is None:
        slot["cnpqCaptacao"] = capt_m
        added += 1
    if slot.get("cnpqVinculos") is None:
        slot["cnpqVinculos"] = vinc_i
        added += 1
    src = sources.setdefault(sigla, {})
    src.setdefault("cnpqVinculos", SRC + " / 'Número de vínculos de fomento do CNPq'")

data["enrichedCnpq"] = datetime.now().isoformat(timespec="seconds")
backup = JSON_PATH.with_suffix(f".backup-cnpq-{datetime.now():%Y%m%d-%H%M%S}.json")
shutil.copy2(JSON_PATH, backup)
JSON_PATH.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
anos = sorted({a for (_, a) in acc})
print(f"OK: {added} campos; anos {anos}. Backup: {backup.name}")
