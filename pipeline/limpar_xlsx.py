"""
Limpa arquivos xlsx com linhas/colunas fantasma.
- Remove colunas cujo cabeçalho é None ou 'Coluna<N>'
- Remove linhas onde TODOS os valores são None
- Para CNPq: filtra apenas linhas com 14_Sigla UF == "PR"
Usa openpyxl em modo read_only (streaming) para evitar OOM.
Saída salva em data/ com prefixo LIMPO_.
"""
import openpyxl
import os, sys, gc, re, time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DATA_DIR = Path(__file__).parent.parent / "data"
PROBE_COLS = 300


def is_real_col(val):
    """False se a coluna é fantasma (None ou padrão 'Coluna<N>')."""
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and not re.match(r"^[Cc]oluna\d+$", s)


def process(src_name, dst_name, filter_col=None, filter_val=None):
    src = DATA_DIR / src_name
    dst = DATA_DIR / dst_name

    src_mb = os.path.getsize(src) / 1e6
    print(f"\n{'='*60}")
    print(f"Arquivo : {src_name}")
    print(f"Tamanho : {src_mb:.1f} MB")

    wb_in  = openpyxl.load_workbook(src, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sname in wb_in.sheetnames:
        t0 = time.time()
        print(f"\n  Aba: '{sname}'")
        ws_in  = wb_in[sname]

        hdr = next(
            ws_in.iter_rows(min_row=1, max_row=1,
                            max_col=PROBE_COLS, values_only=True),
            ()
        )
        real_idxs = [i for i, v in enumerate(hdr) if is_real_col(v)]
        if not real_idxs:
            print(f"    Nenhuma coluna real → aba ignorada.")
            continue

        real_max_col = max(real_idxs) + 1
        print(f"    {len(real_idxs)} colunas reais | max_col={real_max_col}")

        fc_idx = None
        if filter_col:
            for i, v in enumerate(hdr):
                if str(v or "").strip().lower() == filter_col.strip().lower():
                    fc_idx = i
                    break
            status = f"índice {fc_idx}" if fc_idx is not None else "NÃO ENCONTRADA"
            print(f"    Filtro: '{filter_col}' = '{filter_val}' → {status}")

        ws_out   = wb_out.create_sheet(title=sname)
        written  = 0
        skipped  = 0

        ws_out.append([hdr[i] for i in real_idxs])
        written += 1

        for row in ws_in.iter_rows(min_row=2,
                                   max_col=real_max_col,
                                   values_only=True):
            if not any(v is not None for v in row):
                skipped += 1
                continue

            if fc_idx is not None:
                cell_val = row[fc_idx] if fc_idx < len(row) else None
                if str(cell_val or "").strip().upper() != str(filter_val or "").strip().upper():
                    skipped += 1
                    continue

            ws_out.append([row[i] if i < len(row) else None for i in real_idxs])
            written += 1

            if written % 50_000 == 1 and written > 1:
                elapsed = time.time() - t0
                print(f"    ... {written-1:,} linhas escritas ({elapsed:.0f}s)")

        data_rows = written - 1
        print(f"    Resultado: {data_rows:,} linhas × {len(real_idxs)} colunas "
              f"({skipped:,} ignoradas) [{time.time()-t0:.1f}s]")
        gc.collect()

    wb_in.close()
    wb_out.save(dst)
    wb_out.close()
    gc.collect()

    dst_mb = os.path.getsize(dst) / 1e6
    print(f"\n  => Salvo: '{dst_name}'  ({dst_mb:.1f} MB,  "
          f"redução: {src_mb - dst_mb:.1f} MB)")
    return dst_mb


t_start = time.time()

process(
    "Base CAPES- Pós-Graduação - Brasil.xlsx",
    "LIMPO_Base CAPES- Pós-Graduação - Brasil.xlsx",
)

process(
    "Base Cursos - Brasil.xlsx",
    "LIMPO_Base Cursos - Brasil.xlsx",
)

process(
    "Base CNPq - Brasil.xlsx",
    "LIMPO_Base CNPq - Brasil.xlsx",
    filter_col="14_Sigla UF",
    filter_val="PR",
)

print(f"\n\nTodos concluídos em {time.time()-t_start:.0f}s.")
