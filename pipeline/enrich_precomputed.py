#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enriquece data/seti_precomputed.json com dados REAIS que as bases já contêm
mas o assemble_final.py não extrai:

  1. Histórico 2020–2024 da Base Cursos (hoje o painel retroprojetava
     2020–2023 com fatores sintéticos):
     students/entrants/graduates/vacancies/courses/occupancy/dropout/completion
  2. Campos da Base Cursos que alimentavam fórmulas sintéticas na aba 3:
     turno (diurno/noturno), vagas novas/ingresso, escola pública,
     mobilidade acadêmica, vagas não ocupadas
  3. Composição real por grau acadêmico (Bacharelado/Licenciatura/Tecnólogo)
     com desvinculação/conclusão por grau (abas 3 e 4)
  4. Histórico 2020–2024 da Base IES: doutores, docentes estrangeiros (IND-8),
     Portal CAPES (IND-9) e total de docentes (substitui alunos÷15)

As chaves são gravadas em byYear[SIGLA][ano] já com os nomes prefixados que o
painel lê (cursosStudents, cursosOccupancyDay, iesDocDout, ...), de modo que
nenhuma mudança é necessária no loader do data-hub.

Uso:  python3 pipeline/enrich_precomputed.py
"""
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
JSON_PATH = DATA_DIR / "seti_precomputed.json"
YEARS = {2020, 2021, 2022, 2023, 2024}

# ── CO_IES_MAP: extraído do assemble_final.py (fonte única) ──────────────────
_src = (ROOT / "pipeline" / "assemble_final.py").read_text(encoding="utf-8")
_m = re.search(r"CO_IES_MAP = \{(.*?)\n\}", _src, re.S)
CO_IES_MAP = {}
for co, sigla in re.findall(r"(\d+)\s*:\s*\"([A-Za-z]+)\"", _m.group(1)):
    CO_IES_MAP[int(co)] = sigla  # casing canônico (= lista IEES); ver normalize_json_keys.py
print(f"CO_IES_MAP: {len(CO_IES_MAP)} IES")


def r1(v):
    return round(v, 1) if v is not None else None


def r2(v):
    return round(v, 2) if v is not None else None


def pct(num, den, digits=2):
    return round(num / den * 100, digits) if den else None


def safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


# ── 1+2+3. Base Cursos — agregação por IES × ano (+ grau acadêmico) ──────────
print("Lendo Base Cursos - Brasil.xlsx ...", flush=True)
wb = openpyxl.load_workbook(DATA_DIR / "Base Cursos - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["_IES PÚBLICAS ESTADUAIS_CURSOS "]

# Índices 0-based (ver cabeçalho da planilha)
C = dict(ANO=0, MUN=6, CO=15, GRAU=26, CURSO=30, VG=31, VG_DIU=32, VG_NOT=33,
         VG_NOVA=35, ING=47, ING_DIU=50, ING_NOT=51, ING_NOVA=52,
         MAT=77, MAT_DIU=80, MAT_NOT=81, CONC=96, DESV=195,
         ING_PUB=198, ING_PRIV=199, MAT_MOB=221, TIPO=223)
MAXC = max(C.values()) + 1
GRAUS = {"Bacharelado", "Licenciatura", "Tecnólogo"}
GRAU_BY_CODE = {1: "Bacharelado", 2: "Licenciatura", 3: "Tecnólogo"}

agg = {}      # (sigla, ano) -> dict de somas
grau_agg = {} # (sigla, ano, grau) -> dict de somas
SUM_KEYS = ["CURSO", "VG", "VG_DIU", "VG_NOT", "VG_NOVA", "ING", "ING_DIU",
            "ING_NOT", "ING_NOVA", "MAT", "MAT_DIU", "MAT_NOT", "CONC",
            "DESV", "ING_PUB", "ING_PRIV", "MAT_MOB"]

empty_streak = 0
nrows = 0
for row in ws.iter_rows(min_row=2, max_col=MAXC, values_only=True):
    co = row[C["CO"]]
    ano = row[C["ANO"]]
    if co is None and ano is None:
        empty_streak += 1
        if empty_streak > 3000:  # bloco de linhas vazias no fim do arquivo
            break
        continue
    empty_streak = 0
    try:
        co = int(co)
        ano = int(ano)
    except (TypeError, ValueError):
        continue
    if co not in CO_IES_MAP or ano not in YEARS:
        continue
    sigla = CO_IES_MAP[co]
    nrows += 1

    key = (sigla, ano)
    a = agg.setdefault(key, {k: 0 for k in SUM_KEYS})
    for k in SUM_KEYS:
        a[k] += safe_int(row[C[k]])

    grau = row[C["TIPO"]]
    if grau not in GRAUS:
        grau = GRAU_BY_CODE.get(safe_int(row[C["GRAU"]]))
    if grau in GRAUS:
        g = grau_agg.setdefault((sigla, ano, grau), {k: 0 for k in
                                ["VG", "CURSO", "MAT", "ING", "CONC", "DESV"]})
        for k in g:
            g[k] += safe_int(row[C[k]])
wb.close()
print(f"Base Cursos: {nrows} linhas agregadas, {len(agg)} pares IES×ano", flush=True)

cursos_fields = {}
for (sigla, ano), a in agg.items():
    f = {
        # nomes brutos (transparência/depuração)
        "students": a["MAT"], "entrants": a["ING"], "graduates": a["CONC"],
        "vacancies": a["VG"], "courses": a["CURSO"],
        "occupancy": pct(a["ING"], a["VG"]),
        "dropout": pct(a["DESV"], a["MAT"]),
        "completion": pct(a["CONC"], a["MAT"]),
        # nomes prefixados que byYear() do painel consome diretamente
        "cursosStudents": a["MAT"], "cursosEntrants": a["ING"],
        "cursosGraduates": a["CONC"], "cursosVacancies": a["VG"],
        "cursosCourses": a["CURSO"],
        "cursosOccupancy": pct(a["ING"], a["VG"]),
        "cursosDropout": pct(a["DESV"], a["MAT"]),
        "cursosCompletion": pct(a["CONC"], a["MAT"]),
        "cursosVacanciesDay": a["VG_DIU"], "cursosVacanciesNight": a["VG_NOT"],
        "cursosMatDay": a["MAT_DIU"], "cursosMatNight": a["MAT_NOT"],
        "cursosOccupancyDay": pct(a["ING_DIU"], a["VG_DIU"], 1),
        "cursosOccupancyNight": pct(a["ING_NOT"], a["VG_NOT"], 1),
        "cursosVacanciesNova": a["VG_NOVA"],
        "cursosIngressOccupancy": pct(a["ING_NOVA"], a["VG_NOVA"], 1),
        "cursosVacanciesNovaUnfilled": max(a["VG_NOVA"] - a["ING_NOVA"], 0),
        "cursosVacanciesUnfilled": max(a["VG"] - a["ING"], 0),
        "cursosPublicSchool": pct(a["ING_PUB"], a["ING_PUB"] + a["ING_PRIV"], 1),
        "cursosMobility": pct(a["MAT_MOB"], a["MAT"], 2),
    }
    mix = {}
    for grau in GRAUS:
        g = grau_agg.get((sigla, ano, grau))
        if not g:
            continue
        mix[grau] = {
            "vacancies": g["VG"], "courses": g["CURSO"], "students": g["MAT"],
            "entrants": g["ING"], "graduates": g["CONC"],
            "dropout": pct(g["DESV"], g["MAT"]),
            "completion": pct(g["CONC"], g["MAT"]),
        }
    if mix:
        f["grauMix"] = mix
    cursos_fields[(sigla, str(ano))] = f


# ── 4. Base IES — doutores, estrangeiros, Portal CAPES, total docentes ───────
print("Lendo Base IES - Brasil.xlsx ...", flush=True)
wb = openpyxl.load_workbook(DATA_DIR / "Base IES - Brasil.xlsx", read_only=True, data_only=True)
ws = wb["Base_ IES_BRASIL"]
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
hidx = {h: i for i, h in enumerate(headers) if h is not None}
I = dict(ANO=hidx["NU_ANO_CENSO"], CO=hidx["CO_IES"],
         DOC_TOTAL=hidx["QT_DOC_TOTAL"], DOC_EXE=hidx["QT_DOC_EXE"],
         PORTAL=hidx["IN_ACESSO_PORTAL_CAPES"],
         P_DOUT=hidx["Proporção de docentes com doutorado"],
         P_EST=hidx["Proporção de docentes estrangeiros"])

ies_fields = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        co = int(row[I["CO"]])
        ano = int(row[I["ANO"]])
    except (TypeError, ValueError):
        continue
    if co not in CO_IES_MAP or ano not in YEARS:
        continue
    sigla = CO_IES_MAP[co]
    p_dout = row[I["P_DOUT"]]
    p_est = row[I["P_EST"]]
    portal = row[I["PORTAL"]]
    f = {}
    if p_dout is not None:
        f["iesDocDout"] = r2(float(p_dout) * 100)
        f["doctors"] = f["iesDocDout"]
    if p_est is not None:
        f["iesDocForeign"] = r2(float(p_est) * 100)
    if portal is not None:
        f["iesCapesPortal"] = 1 if str(portal).strip() in ("1", "1.0", "True", "Sim") else 0
    if row[I["DOC_TOTAL"]] is not None:
        f["docTotal"] = safe_int(row[I["DOC_TOTAL"]])
    if row[I["DOC_EXE"]] is not None:
        f["docExe"] = safe_int(row[I["DOC_EXE"]])
    if f:
        ies_fields[(sigla, str(ano))] = f
wb.close()
print(f"Base IES: {len(ies_fields)} pares IES×ano", flush=True)


# ── Mescla no JSON (sem sobrescrever campos existentes) ──────────────────────
data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
by_year = data.setdefault("byYear", {})
sources = data.setdefault("sources", {})

SRC_CURSOS = "Base Cursos - Brasil.xlsx / _IES PÚBLICAS ESTADUAIS_CURSOS / soma por CO_IES e ano (2020–2024)"
SRC_IES = "Base IES - Brasil.xlsx / Base_ IES_BRASIL / por CO_IES e ano (2020–2024)"

added = 0
for (sigla, ano), f in cursos_fields.items():
    rec = by_year.setdefault(sigla, {}).setdefault(ano, {})
    for k, v in f.items():
        if v is None:
            continue
        if k not in rec or rec[k] is None:
            rec[k] = v
            added += 1
for (sigla, ano), f in ies_fields.items():
    rec = by_year.setdefault(sigla, {}).setdefault(ano, {})
    for k, v in f.items():
        if v is None:
            continue
        if k not in rec or rec[k] is None:
            rec[k] = v
            added += 1

for sigla in {s for (s, _) in cursos_fields}:
    src = sources.setdefault(sigla, {})
    src.setdefault("historico_2020_2024", SRC_CURSOS)
    src.setdefault("cursosOccupancyDay", SRC_CURSOS + " / QT_ING_DIURNO ÷ QT_VG_TOTAL_DIURNO × 100")
    src.setdefault("cursosOccupancyNight", SRC_CURSOS + " / QT_ING_NOTURNO ÷ QT_VG_TOTAL_NOTURNO × 100")
    src.setdefault("cursosPublicSchool", SRC_CURSOS + " / QT_ING_PROCESCPUBLICA ÷ (PUB+PRIV) × 100")
    src.setdefault("cursosMobility", SRC_CURSOS + " / QT_MAT_MOB_ACADEMICA ÷ QT_MAT × 100")
    src.setdefault("cursosIngressOccupancy", SRC_CURSOS + " / QT_ING_VG_NOVA ÷ QT_VG_NOVA × 100")
    src.setdefault("grauMix", SRC_CURSOS + " / agregado por tipo_curso (Bacharelado/Licenciatura/Tecnólogo)")
for sigla in {s for (s, _) in ies_fields}:
    src = sources.setdefault(sigla, {})
    src.setdefault("iesDocForeign", SRC_IES + " / 'Proporção de docentes estrangeiros' × 100")
    src.setdefault("iesCapesPortal", SRC_IES + " / IN_ACESSO_PORTAL_CAPES")
    src.setdefault("docTotal", SRC_IES + " / QT_DOC_TOTAL")

data["enriched"] = datetime.now().isoformat(timespec="seconds")

backup = JSON_PATH.with_suffix(f".backup-{datetime.now():%Y%m%d-%H%M%S}.json")
shutil.copy2(JSON_PATH, backup)
JSON_PATH.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
print(f"OK: {added} campos adicionados. Backup: {backup.name}")
print(f"JSON: {JSON_PATH} ({JSON_PATH.stat().st_size / 1e6:.1f} MB)")
