import openpyxl
import os
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"

files = [f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx")]

for fname in sorted(files):
    path = DATA_DIR / fname
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print(f"\n=== {fname} ===")
        print(f"Sheets: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
            print(f"  Sheet [{sname}]:")
            for r in rows:
                row_str = str(r)
                if len(row_str) > 300:
                    row_str = row_str[:300] + "..."
                print(f"    {row_str}")
        wb.close()
    except Exception as e:
        print(f"ERRO em {fname}: {e}")
