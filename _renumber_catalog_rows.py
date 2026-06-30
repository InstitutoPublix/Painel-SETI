# -*- coding: utf-8 -*-
import openpyxl

PATH = "data/5. Relação de Indicadores das Universidades_v2.xlsx"

wb = openpyxl.load_workbook(PATH)
ws = wb["Indicadores atualizados"]

# Sheet rows 96-109 currently hold Código 95-108 (in order).
# Renumber to 109-122, same order, no other column touched.
OLD_START_ROW = 96
OLD_START_CODE = 95
NEW_START_CODE = 109
COUNT = 14

changes = []
for i in range(COUNT):
    row = OLD_START_ROW + i
    old_code = ws.cell(row=row, column=1).value
    expected_old = OLD_START_CODE + i
    assert old_code == expected_old, f"row {row}: expected código {expected_old}, found {old_code!r}"
    new_code = NEW_START_CODE + i
    nome = ws.cell(row=row, column=2).value
    ws.cell(row=row, column=1, value=new_code)
    changes.append((row, old_code, new_code, nome))

wb.save(PATH)

for row, old_code, new_code, nome in changes:
    print(f"row {row}: {old_code} -> {new_code}  ({nome})")
print("Saved:", PATH)
